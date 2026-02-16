#!/usr/bin/env python3
"""Test v6.0 changes: SMTP email, verified links, Excel with charts."""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from hands.flight_search import (
    _generate_flight_excel, _verify_booking_link, _get_verified_booking_link,
    _get_airline_booking_url, _html_flight_report_email, _send_html_email,
    _get_smtp_password, _send_html_email_mailapp
)

print("=" * 60)
print("  TARS Flight Engine v6.0 — Test Suite")
print("=" * 60)

# ── Test 1: Airline Deep Links ──
print("\n📌 TEST 1: Airline Deep Links (previously broken)")
test_airlines = ['Emirates', 'Turkish Airlines', 'PIA', 'Singapore Airlines', 'Etihad', 'Korean Air']
for airline in test_airlines:
    url = _get_airline_booking_url(airline, 'TPA', 'ISB', '2025-03-15', '2025-03-25')
    has_params = '?' in url and '=' in url
    status = "✅ deep link" if has_params else "⚠️ generic"
    print(f"  {airline:25s}: {status} → {url[:65]}...")

# ── Test 2: SMTP password loader ──
print("\n📌 TEST 2: SMTP Password Config")
pwd = _get_smtp_password()
print(f"  SMTP password configured: {'✅ Yes' if pwd else '⚠️ No (will fallback to Mail.app)'}")

# ── Test 3: Excel with Charts ──
print("\n📌 TEST 3: Excel Dashboard with Charts + Conditional Formatting")
flights = [
    {'price': '$59', 'airline': 'Spirit', 'depart_time': '6:00 AM', 'arrive_time': '9:30 AM',
     'duration': '3h 30m', 'stops': 'Nonstop', 'value_score': 95, 'value_label': '🟢 Excellent',
     'booking_link': 'https://www.spirit.com', 'fare_class': 'Basic Economy', 'baggage': 'No carry-on'},
    {'price': '$89', 'airline': 'Frontier', 'depart_time': '7:15 AM', 'arrive_time': '10:45 AM',
     'duration': '3h 30m', 'stops': 'Nonstop', 'value_score': 82, 'value_label': '🟢 Excellent',
     'booking_link': 'https://www.flyfrontier.com', 'fare_class': 'Economy', 'baggage': 'Carry-on included'},
    {'price': '$142', 'airline': 'Delta', 'depart_time': '8:00 AM', 'arrive_time': '11:30 AM',
     'duration': '3h 30m', 'stops': 'Nonstop', 'value_score': 75, 'value_label': '🔵 Great',
     'booking_link': 'https://www.delta.com', 'fare_class': 'Main Cabin', 'baggage': 'Checked bag included'},
    {'price': '$178', 'airline': 'United', 'depart_time': '10:00 AM', 'arrive_time': '2:30 PM',
     'duration': '4h 30m', 'stops': '1 stop', 'value_score': 55, 'value_label': '🟡 Good',
     'booking_link': 'https://www.united.com', 'layover_airport': 'ATL', 'layover_duration': '1h 15m'},
    {'price': '$225', 'airline': 'American', 'depart_time': '12:00 PM', 'arrive_time': '5:00 PM',
     'duration': '5h', 'stops': '1 stop', 'value_score': 40, 'value_label': '🟠 Fair',
     'booking_link': 'https://www.aa.com', 'layover_airport': 'DFW', 'layover_duration': '2h'},
]

analytics = {
    'price_min': 59, 'price_max': 225, 'price_avg': 139, 'price_median': 142,
    'price_range': 166, 'price_std_dev': 60, 'total_flights': 5,
    'nonstop_count': 3, 'connecting_count': 2, 'nonstop_premium': 0,
    'airline_count': 5,
    'airline_stats': {
        'Spirit': {'min': 59, 'avg': 59, 'count': 1, 'nonstop': 1},
        'Frontier': {'min': 89, 'avg': 89, 'count': 1, 'nonstop': 1},
        'Delta': {'min': 142, 'avg': 142, 'count': 1, 'nonstop': 1},
        'United': {'min': 178, 'avg': 178, 'count': 1, 'nonstop': 0},
        'American': {'min': 225, 'avg': 225, 'count': 1, 'nonstop': 0},
    }
}

suggestions = [
    {'type': 'nonstop_value', 'icon': '✈️', 'text': 'Nonstop available at $59 — great deal!', 'priority': 1},
    {'type': 'price_spread', 'icon': '📊', 'text': 'Price range is $166. Picking wisely saves up to 73%.', 'priority': 1},
    {'type': 'best_value', 'icon': '⭐', 'text': 'Best value: Spirit at $59 (Nonstop, 3h 30m) — Score 95/100.', 'priority': 1},
]

summary = {
    'Route': 'TPA → JFK',
    'Departure': '2025-03-15',
    'Total Options': '5',
    'Cheapest': '$59 — Spirit (Nonstop)',
}

result = _generate_flight_excel('Flights TPA→JFK 2025-03-15', flights, 'TPA', 'JFK',
                                 'https://google.com/travel/flights', summary,
                                 analytics=analytics, suggestions=suggestions)

if result.get('success'):
    size = os.path.getsize(result['path'])
    print(f"  ✅ Excel generated: {result['path']}")
    print(f"  📊 File size: {size:,} bytes")
    
    # Check it has multiple sheets
    from openpyxl import load_workbook
    wb = load_workbook(result['path'])
    print(f"  📑 Sheets: {wb.sheetnames}")
    
    # Check for charts
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        charts = ws._charts if hasattr(ws, '_charts') else []
        cf_count = len(ws.conditional_formatting._cf_rules) if hasattr(ws.conditional_formatting, '_cf_rules') else 0
        print(f"    {sheet_name}: {len(charts)} charts, {cf_count} conditional format rules")
    wb.close()
else:
    print(f"  ❌ Excel failed: {result}")

# ── Test 4: HTML Email Generation ──
print("\n📌 TEST 4: HTML Email Generation")
html = _html_flight_report_email('TPA', 'JFK', '2025-03-15', '2025-03-25', flights,
                                  'https://google.com/travel/flights',
                                  price_insight="Prices are currently low for this route",
                                  tracker_suggestion="💡 Set a price tracker at $50")
print(f"  HTML length: {len(html):,} chars")
print(f"  Has chart bars: {'✅' if 'Price Comparison' in html else '❌'}")
print(f"  Has value badges: {'✅' if 'CHEAPEST' in html else '❌'}")
print(f"  Has suggestions: {'✅' if 'Smart Suggestions' in html else '❌'}")
print(f"  Has price insight: {'✅' if 'Google Insight' in html else '❌'}")
print(f"  Has v6 branding: {'✅' if 'v6' in html else '❌'}")
print(f"  Has verified note: {'✅' if 'verified' in html.lower() else '❌'}")

# ── Test 5: Link Verification ──
print("\n📌 TEST 5: Booking Link Verification")
test_links = [
    ("Google Flights", "https://www.google.com/travel/flights"),
    ("Delta", "https://www.delta.com"),
    ("Fake/Broken", "https://www.definitely-not-a-real-airline-website-xyz.com"),
]
for name, url in test_links:
    ok = _verify_booking_link(url)
    print(f"  {name:20s}: {'✅ reachable' if ok else '❌ broken'}")

print("\n" + "=" * 60)
print("  v6.0 Test Suite Complete!")
print("=" * 60)
