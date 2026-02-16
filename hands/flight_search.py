"""
╔══════════════════════════════════════════════════════════════╗
║       TARS — Flight Engine v6.0 (Visual + Verified)          ║
╠══════════════════════════════════════════════════════════════╣
║                                                              ║
║  Phase 1:  Core Engine — Google Flights URL builder, parser  ║
║  Phase 2:  Booking Links — Google Flights link per flight    ║
║  Phase 2B: Airline Direct — 40+ airline deep link generator  ║
║  Phase 3:  Smart Sampling — 6-month range (weekly cadence)   ║
║  Phase 4:  HTML Email Templates — airline-grade design       ║
║  Phase 5:  Enhanced Excel — charts, data bars, dashboard     ║
║  Phase 6:  Price Tracker — persistent JSON DB for routes     ║
║  Phase 6B: SMTP Email — direct Outlook SMTP, no Mail.app     ║
║  Phase 7:  Alert Engine — background price checks + alerts   ║
║  Phase 8:  Rich HTML Email — properly rendered via SMTP      ║
║  Phase 9:  Pipeline Integration — report + email + notify    ║
║  Phase 10: Tool Registration — brain tools + executor        ║
║  Phase 11: Intelligence Layer — suggestions, analytics,      ║
║            nearby airports, value scores, price insights     ║
║  Phase 12: Structured DOM Parser — JSON from DOM, not regex  ║
║  Phase 13: Return Flight + Layover + Fare + Baggage parsing  ║
║  Phase 14: Google Price Insights scraping                    ║
║  Phase 15: Search Cache + CDP Retry + Parallel Scanning      ║
║  Phase 16: Link Verification — HTTP check before sending     ║
║  Phase 17: Excel Dashboard — charts, conditional formatting, ║
║            KPI cards, bar/pie charts, color scales            ║
║                                                              ║
║  ⚠️ ONLY Google Flights. NEVER Kayak/Skyscanner/Expedia.    ║
╚══════════════════════════════════════════════════════════════╝
"""

import json
import math
import os
import re
import hashlib
import threading
import time
import urllib.parse
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

from hands.cdp import CDP


# ═══════════════════════════════════════════════════════
#  PHASE 15A — Search Cache (15-min TTL)
# ═══════════════════════════════════════════════════════

_search_cache = {}  # key → {"result": ..., "ts": float}
_CACHE_TTL = 900  # 15 minutes


def _cache_key(origin, destination, depart_date, return_date, cabin, stops):
    """Generate a unique cache key for a flight search."""
    raw = f"{origin}|{destination}|{depart_date}|{return_date}|{cabin}|{stops}"
    return hashlib.md5(raw.encode()).hexdigest()


def _get_cached(key):
    """Get a cached result if still valid."""
    entry = _search_cache.get(key)
    if entry and (time.time() - entry["ts"]) < _CACHE_TTL:
        return entry["result"]
    return None


def _set_cache(key, result):
    """Cache a search result."""
    _search_cache[key] = {"result": result, "ts": time.time()}
    # Evict old entries (keep max 50)
    if len(_search_cache) > 50:
        oldest = sorted(_search_cache, key=lambda k: _search_cache[k]["ts"])
        for old_key in oldest[:10]:
            _search_cache.pop(old_key, None)


def _load_config():
    """Load TARS config for iMessage sender."""
    import yaml
    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config.yaml")
    try:
        with open(config_path) as f:
            return yaml.safe_load(f)
    except Exception:
        return {"imessage": {"owner_phone": "", "rate_limit": 3, "max_message_length": 1600}}


class _FlightSearchTimeout(Exception):
    """Raised when flight search exceeds time limit."""
    pass


# ═══════════════════════════════════════════════════════
#  PHASE 1 — Airport Code Lookup
# ═══════════════════════════════════════════════════════

CITY_TO_IATA = {
    # ── US Major ──
    "new york": "JFK", "nyc": "JFK", "jfk": "JFK", "laguardia": "LGA", "newark": "EWR",
    "los angeles": "LAX", "la": "LAX", "lax": "LAX",
    "chicago": "ORD", "ohare": "ORD", "midway": "MDW",
    "san francisco": "SFO", "sf": "SFO",
    "miami": "MIA",
    "dallas": "DFW", "dfw": "DFW",
    "houston": "IAH", "hobby": "HOU",
    "atlanta": "ATL",
    "boston": "BOS",
    "seattle": "SEA",
    "denver": "DEN",
    "las vegas": "LAS", "vegas": "LAS",
    "orlando": "MCO",
    "tampa": "TPA",
    "phoenix": "PHX",
    "detroit": "DTW",
    "minneapolis": "MSP",
    "philadelphia": "PHL", "philly": "PHL",
    "washington": "IAD", "dc": "DCA", "dulles": "IAD", "reagan": "DCA",
    "san diego": "SAN",
    "austin": "AUS",
    "nashville": "BNA",
    "portland": "PDX",
    "charlotte": "CLT",
    "salt lake city": "SLC", "slc": "SLC",
    "pittsburgh": "PIT",
    "st louis": "STL", "saint louis": "STL",
    "honolulu": "HNL", "hawaii": "HNL",
    # ── US Additional ──
    "jacksonville": "JAX", "raleigh": "RDU", "raleigh durham": "RDU",
    "indianapolis": "IND", "kansas city": "MCI", "columbus": "CMH",
    "cincinnati": "CVG", "cleveland": "CLE", "milwaukee": "MKE",
    "new orleans": "MSY", "memphis": "MEM", "richmond": "RIC",
    "norfolk": "ORF", "birmingham": "BHM", "louisville": "SDF",
    "buffalo": "BUF", "rochester": "ROC", "syracuse": "SYR",
    "albany": "ALB", "hartford": "BDL", "sacramento": "SMF",
    "san antonio": "SAT", "el paso": "ELP", "tucson": "TUS",
    "albuquerque": "ABQ", "omaha": "OMA", "oklahoma city": "OKC",
    "tulsa": "TUL", "boise": "BOI", "spokane": "GEG",
    "anchorage": "ANC", "maui": "OGG", "kona": "KOA",
    "savannah": "SAV", "charleston": "CHS", "myrtle beach": "MYR",
    "key west": "EYW", "fort myers": "RSW", "west palm beach": "PBI",
    "fort lauderdale": "FLL", "st pete": "PIE", "sarasota": "SRQ",
    "daytona beach": "DAB", "pensacola": "PNS", "tallahassee": "TLH",
    "burbank": "BUR", "long beach": "LGB", "ontario": "ONT",
    "orange county": "SNA", "palm springs": "PSP", "santa barbara": "SBA",
    "oakland": "OAK", "san jose": "SJC", "reno": "RNO",
    "colorado springs": "COS", "little rock": "LIT", "des moines": "DSM",
    "grand rapids": "GRR", "providence": "PVD", "manchester": "MHT",
    "baltimore": "BWI",
    # ── International ──
    "london": "LHR", "heathrow": "LHR", "gatwick": "LGW",
    "paris": "CDG",
    "tokyo": "NRT", "narita": "NRT", "haneda": "HND",
    "dubai": "DXB",
    "singapore": "SIN",
    "hong kong": "HKG",
    "toronto": "YYZ",
    "vancouver": "YVR",
    "sydney": "SYD",
    "melbourne": "MEL",
    "mumbai": "BOM", "bombay": "BOM",
    "delhi": "DEL", "new delhi": "DEL",
    "bangkok": "BKK",
    "istanbul": "IST",
    "rome": "FCO",
    "madrid": "MAD",
    "barcelona": "BCN",
    "amsterdam": "AMS",
    "frankfurt": "FRA",
    "berlin": "BER",
    "munich": "MUC",
    "zurich": "ZRH",
    "cancun": "CUN",
    "mexico city": "MEX",
    "sao paulo": "GRU",
    "rio": "GIG", "rio de janeiro": "GIG",
    "cairo": "CAI",
    "doha": "DOH",
    "seoul": "ICN", "incheon": "ICN",
    "beijing": "PEK",
    "shanghai": "PVG",
    "kuala lumpur": "KUL",
    "jakarta": "CGK",
    "johannesburg": "JNB",
    "nairobi": "NBO",
    "buenos aires": "EZE",
    "bogota": "BOG",
    "lima": "LIM",
    "lisbon": "LIS",
    "dublin": "DUB",
    "edinburgh": "EDI",
    "manchester uk": "MAN",
    "copenhagen": "CPH",
    "stockholm": "ARN",
    "oslo": "OSL",
    "helsinki": "HEL",
    "vienna": "VIE",
    "prague": "PRG",
    "warsaw": "WAW",
    "athens": "ATH",
    "milan": "MXP",
    # ── South Asia ──
    "kathmandu": "KTM", "nepal": "KTM",
    "colombo": "CMB", "sri lanka": "CMB",
    "dhaka": "DAC", "bangladesh": "DAC",
    "islamabad": "ISB", "lahore": "LHE", "karachi": "KHI",
    "peshawar": "PEW", "faisalabad": "LYP", "multan": "MUX",
    "sialkot": "SKT", "quetta": "UET",
    "chennai": "MAA", "madras": "MAA",
    "bangalore": "BLR", "bengaluru": "BLR",
    "hyderabad": "HYD",
    "kolkata": "CCU", "calcutta": "CCU",
    "goa": "GOI",
    "ahmedabad": "AMD", "pune": "PNQ", "jaipur": "JAI",
    "lucknow": "LKO", "cochin": "COK", "kochi": "COK",
    "trivandrum": "TRV", "amritsar": "ATQ",
    # ── Southeast Asia ──
    "manila": "MNL", "philippines": "MNL",
    "hanoi": "HAN", "ho chi minh": "SGN", "saigon": "SGN",
    "phnom penh": "PNH", "cambodia": "PNH",
    "bali": "DPS",
    "phuket": "HKT",
    "taipei": "TPE", "taiwan": "TPE",
    "chiang mai": "CNX", "cebu": "CEB", "yangon": "RGN",
    # ── Middle East / Africa ──
    "abu dhabi": "AUH",
    "riyadh": "RUH", "jeddah": "JED", "saudi": "RUH",
    "medina": "MED", "dammam": "DMM",
    "muscat": "MCT", "oman": "MCT",
    "bahrain": "BAH", "kuwait": "KWI",
    "tel aviv": "TLV", "israel": "TLV",
    "amman": "AMM", "jordan": "AMM",
    "beirut": "BEY", "lebanon": "BEY",
    "baghdad": "BGW", "iraq": "BGW", "erbil": "EBL",
    "addis ababa": "ADD", "ethiopia": "ADD",
    "lagos": "LOS", "nigeria": "LOS", "abuja": "ABV",
    "casablanca": "CMN", "morocco": "CMN", "marrakech": "RAK",
    "cape town": "CPT",
    "dar es salaam": "DAR", "tanzania": "DAR",
    "accra": "ACC", "ghana": "ACC",
    "tunis": "TUN", "tunisia": "TUN",
    "algiers": "ALG",
    # ── Americas ──
    "havana": "HAV", "cuba": "HAV",
    "san jose": "SJC", "costa rica": "SJO",
    "panama city": "PTY", "panama": "PTY",
    "santiago": "SCL", "chile": "SCL",
    "medellin": "MDE",
    "quito": "UIO", "ecuador": "UIO",
    "montevideo": "MVD", "uruguay": "MVD",
    "punta cana": "PUJ", "santo domingo": "SDQ",
    "kingston": "KIN", "jamaica": "KIN",
    "nassau": "NAS", "bahamas": "NAS",
    "calgary": "YYC", "montreal": "YUL", "ottawa": "YOW",
    "edmonton": "YEG", "winnipeg": "YWG", "halifax": "YHZ",
    # ── Oceania ──
    "auckland": "AKL", "new zealand": "AKL",
    "wellington": "WLG",
    "brisbane": "BNE",
    "perth": "PER",
    "fiji": "NAN",
    "queenstown": "ZQN", "christchurch": "CHC",
    "gold coast": "OOL", "cairns": "CNS",
    # ── East Asia ──
    "osaka": "KIX", "nagoya": "NGO", "fukuoka": "FUK", "sapporo": "CTS",
    "guangzhou": "CAN", "shenzhen": "SZX", "chengdu": "CTU",
    "hangzhou": "HGH", "nanjing": "NKG", "xi'an": "XIY", "xian": "XIY",
    "busan": "PUS",
    "macau": "MFM", "ulaanbaatar": "ULN", "mongolia": "ULN",
    # ── Europe Additional ──
    "nice": "NCE", "lyon": "LYS", "marseille": "MRS",
    "venice": "VCE", "florence": "FLR", "naples": "NAP",
    "porto": "OPO", "seville": "SVQ", "malaga": "AGP",
    "brussels": "BRU", "geneva": "GVA",
    "bucharest": "OTP", "budapest": "BUD",
    "zagreb": "ZAG", "belgrade": "BEG", "sofia": "SOF",
    "split": "SPU", "dubrovnik": "DBV",
    "riga": "RIX", "tallinn": "TLL", "vilnius": "VNO",
    "reykjavik": "KEF", "iceland": "KEF",
    "malta": "MLA", "cyprus": "LCA",
    "santorini": "JTR", "mykonos": "JMK", "crete": "HER",
    "tenerife": "TFS", "canary islands": "TFS",
}


def _resolve_airport(city_or_code: str) -> str:
    """Resolve a city name or airport code to IATA code.
    
    Handles: exact match, 3-letter codes, multi-word inputs like
    'Lahore Pakistan', 'New York City', 'Salt Lake', etc.
    Uses longest-match-first to avoid 'la' matching before 'lahore'.
    """
    raw = city_or_code.strip()
    text = raw.lower()

    # 1. Exact match
    if text in CITY_TO_IATA:
        return CITY_TO_IATA[text]

    # 2. Already a 3-letter IATA code
    if len(text) == 3 and text.isalpha():
        return text.upper()

    # 3. Strip common country/region suffixes and retry exact match
    _STRIP = ['pakistan', 'india', 'china', 'japan', 'korea', 'nepal',
              'bangladesh', 'sri lanka', 'thailand', 'vietnam', 'indonesia',
              'malaysia', 'philippines', 'turkey', 'uae', 'saudi arabia',
              'qatar', 'uk', 'england', 'france', 'germany', 'italy',
              'spain', 'canada', 'mexico', 'brazil', 'australia',
              'usa', 'us', 'united states', 'united kingdom']
    cleaned = text
    for suffix in _STRIP:
        cleaned = cleaned.replace(suffix, '').strip().rstrip(',')
    cleaned = cleaned.strip()
    if cleaned and cleaned in CITY_TO_IATA:
        return CITY_TO_IATA[cleaned]

    # 4. Try each word in the input as an exact city match (longest words first)
    words = text.replace(',', ' ').split()
    # Try multi-word combos first ("salt lake", "new york", "los angeles")
    for n in range(min(3, len(words)), 0, -1):
        for i in range(len(words) - n + 1):
            phrase = ' '.join(words[i:i+n])
            if phrase in CITY_TO_IATA:
                return CITY_TO_IATA[phrase]

    # 5. Substring matching — longest match wins to avoid 'la' beating 'lahore'
    best_match = None
    best_len = 0
    for city, code in CITY_TO_IATA.items():
        if city in text and len(city) > best_len:
            best_match = code
            best_len = len(city)
    if best_match:
        return best_match

    # 6. Final fallback — check if input looks like an IATA code
    upper = raw.upper()
    if len(upper) == 3 and upper.isalpha():
        return upper
    return raw


def _parse_date(date_str: str) -> str:
    """Parse a human date string into YYYY-MM-DD format."""
    if not date_str:
        return ""
    text = date_str.strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}$', text):
        return text
    now = datetime.now()
    formats = [
        "%B %d, %Y", "%B %d %Y", "%B %d",
        "%b %d, %Y", "%b %d %Y", "%b %d",
        "%m/%d/%Y", "%m/%d", "%m-%d-%Y", "%Y/%m/%d",
    ]
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if "%Y" not in fmt and "%y" not in fmt:
                parsed = parsed.replace(year=now.year)
                if parsed < now:
                    parsed = parsed.replace(year=now.year + 1)
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue
    lower = text.lower()
    if "today" in lower:
        return now.strftime("%Y-%m-%d")
    if "tomorrow" in lower:
        return (now + timedelta(days=1)).strftime("%Y-%m-%d")
    if "next week" in lower:
        return (now + timedelta(days=7)).strftime("%Y-%m-%d")
    if "next month" in lower and not re.search(r'\d+\s*month', lower):
        return (now + timedelta(days=30)).strftime("%Y-%m-%d")
    if "this week" in lower:
        return now.strftime("%Y-%m-%d")
    if "this month" in lower:
        return (now + timedelta(days=1)).strftime("%Y-%m-%d")

    # Relative phrases: "in 6 months", "6 months from now", "next 3 months",
    # "in 2 weeks", "3 weeks from now", "in 10 days"
    _WORD_TO_NUM = {'one': 1, 'two': 2, 'three': 3, 'four': 4, 'five': 5,
                    'six': 6, 'seven': 7, 'eight': 8, 'nine': 9, 'ten': 10,
                    'eleven': 11, 'twelve': 12}
    rel = re.search(r'(?:in|next|after)\s+(\d+|' + '|'.join(_WORD_TO_NUM.keys()) + r')\s*(month|week|day)', lower)
    if not rel:
        rel = re.search(r'(\d+|' + '|'.join(_WORD_TO_NUM.keys()) + r')\s*(month|week|day)s?\s*(?:from\s*now|later|ahead|out)', lower)
    if rel:
        num_str, unit = rel.group(1), rel.group(2)
        num = _WORD_TO_NUM.get(num_str, None)
        if num is None:
            try:
                num = int(num_str)
            except ValueError:
                num = 1
        if 'month' in unit:
            result_date = now + timedelta(days=num * 30)
        elif 'week' in unit:
            result_date = now + timedelta(weeks=num)
        else:
            result_date = now + timedelta(days=num)
        return result_date.strftime("%Y-%m-%d")

    import calendar
    for month_num in range(1, 13):
        month_name = calendar.month_name[month_num].lower()
        month_abbr = calendar.month_abbr[month_num].lower()
        if month_name in lower or month_abbr in lower:
            year = now.year
            target = datetime(year, month_num, 15)
            if target < now:
                target = target.replace(year=year + 1)
            return target.strftime("%Y-%m-%d")
    logger.warning(f"    ⚠️ Could not parse date '{text}', defaulting to tomorrow")
    return (now + timedelta(days=1)).strftime("%Y-%m-%d")


# ═══════════════════════════════════════════════════════
#  PHASE 2 — Google Flights URL Builder + Booking Links
# ═══════════════════════════════════════════════════════

def _build_google_flights_url(
    origin: str, destination: str, depart_date: str,
    return_date: str = "", passengers: int = 1,
    trip_type: str = "round_trip", cabin: str = "economy",
    stops: str = "any", sort_by: str = "price",
) -> str:
    """Construct a Google Flights search URL with all parameters encoded."""
    origin_code = _resolve_airport(origin)
    dest_code = _resolve_airport(destination)
    depart = _parse_date(depart_date)
    query_parts = [f"flights from {origin_code} to {dest_code}"]
    if depart:
        query_parts.append(f"on {depart}")
    if trip_type == "one_way":
        query_parts.append("one way")
    elif return_date:
        ret = _parse_date(return_date)
        query_parts.append(f"return {ret}")
    if cabin != "economy":
        query_parts.append(cabin.replace("_", " "))
    if stops == "nonstop":
        query_parts.append("nonstop")
    elif stops == "1stop":
        query_parts.append("1 stop or fewer")
    if passengers > 1:
        query_parts.append(f"{passengers} passengers")
    query = " ".join(query_parts)
    return f"https://www.google.com/travel/flights?q={urllib.parse.quote_plus(query)}"


def _build_booking_link(origin: str, destination: str, depart_date: str,
                        return_date: str = "", trip_type: str = "round_trip") -> str:
    """Build a clean Google Flights booking link for a specific date.
    
    Uses the proper Google Flights URL structure that opens directly to search results.
    Format: /travel/flights/booking?hl=en&gl=us&tfs=...
    Fallback: /travel/flights?q=... (simpler but still works)
    """
    origin_code = _resolve_airport(origin)
    dest_code = _resolve_airport(destination)
    depart = _parse_date(depart_date) if not re.match(r'^\d{4}-\d{2}-\d{2}$', depart_date) else depart_date
    
    # Build proper Google Flights search URL
    # This format pre-fills the search form correctly
    params = {
        "hl": "en",
        "gl": "us", 
        "curr": "USD",
    }
    
    if trip_type == "one_way":
        # One-way: /travel/flights?q=flights+SLC+to+LHE+on+2026-03-20+one+way
        q = f"flights {origin_code} to {dest_code} on {depart} one way"
    elif return_date:
        ret = _parse_date(return_date) if not re.match(r'^\d{4}-\d{2}-\d{2}$', return_date) else return_date
        q = f"flights {origin_code} to {dest_code} departing {depart} returning {ret}"
    else:
        q = f"flights {origin_code} to {dest_code} on {depart}"
    
    params["q"] = q
    return f"https://www.google.com/travel/flights?{urllib.parse.urlencode(params)}"


# ═══════════════════════════════════════════════════════
#  PHASE 2B — Airline Direct Booking URL Generator
# ═══════════════════════════════════════════════════════

# Mapping of airline → their flight search/booking URL templates
# These deep links go directly to each airline's booking page with route pre-filled
AIRLINE_BOOKING_URLS = {
    "delta": {
        "base": "https://www.delta.com/flight-search/book-a-flight",
        "search": lambda o, d, dep, ret: (
            f"https://www.delta.com/flight-search/book-a-flight?cacheKeySuffix=a"
            f"&departureDate={dep}&returnDate={ret or dep}"
            f"&originCity={o}&destinationCity={d}&paxCount=1&searchByCabin=true&cabinSelect=MAIN"
        ),
    },
    "united": {
        "base": "https://www.united.com/en/us",
        "search": lambda o, d, dep, ret: (
            f"https://www.united.com/ual/en/us/flight-search/book-a-flight/results/rev"
            f"?f={o}&t={d}&d={dep}&r={ret or dep}&sc=7&px=1&taxng=1&newHP=True&clm=7&st=bestmatches"
        ),
    },
    "american": {
        "base": "https://www.aa.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.aa.com/booking/search?locale=en_US"
            f"&pax=1&adult=1&type={'roundTrip' if ret else 'oneWay'}"
            f"&searchType=revenue&origin={o}&destination={d}"
            f"&departDate={dep}" + (f"&returnDate={ret}" if ret else "")
        ),
    },
    "southwest": {
        "base": "https://www.southwest.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.southwest.com/air/booking/select.html"
            f"?originationAirportCode={o}&destinationAirportCode={d}"
            f"&departureDate={dep}" + (f"&returnDate={ret}" if ret else "")
            + f"&adultPassengersCount=1&tripType={'roundtrip' if ret else 'oneway'}"
        ),
    },
    "jetblue": {
        "base": "https://www.jetblue.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.jetblue.com/booking/flights"
            f"?from={o}&to={d}&depart={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&pax=1&isMultiCity=false&isRoundTrip={'true' if ret else 'false'}"
        ),
    },
    "spirit": {
        "base": "https://www.spirit.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.spirit.com/book/flights"
            f"?orgCode={o}&desCode={d}&departDate={dep}"
            + (f"&returnDate={ret}" if ret else "")
            + f"&adults=1&tripType={'RT' if ret else 'OW'}"
        ),
    },
    "frontier": {
        "base": "https://www.flyfrontier.com",
        "search": lambda o, d, dep, ret: (
            f"https://booking.flyfrontier.com/Flight/Select"
            f"?o1={o}&d1={d}&dd1={dep}"
            + (f"&dd2={ret}" if ret else "")
            + f"&ADT=1&mon=true"
        ),
    },
    "alaska": {
        "base": "https://www.alaskaair.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.alaskaair.com/shopping/flights"
            f"?A=1&prior=form&ShoppingMethod=Traditional"
            f"&FO={o}&FD={d}&FDD={dep}"
            + (f"&FRD={ret}" if ret else "")
            + f"&RT={'true' if ret else 'false'}"
        ),
    },
    "hawaiian": {
        "base": "https://www.hawaiianairlines.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.hawaiianairlines.com/book/flights"
        ),
    },
    "air canada": {
        "base": "https://www.aircanada.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.aircanada.com/booking/search"
            f"?org0={o}&dest0={d}&departureDate0={dep}"
            + (f"&org1={d}&dest1={o}&departureDate1={ret}" if ret else "")
            + f"&ADT=1&tripType={'RT' if ret else 'OW'}&lang=en-CA"
        ),
    },
    "british airways": {
        "base": "https://www.britishairways.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.britishairways.com/travel/book/public/en_us"
            f"#/flightList?origin={o}&destination={d}&adt=1"
            f"&departureDate={dep}" + (f"&returnDate={ret}" if ret else "")
        ),
    },
    "lufthansa": {
        "base": "https://www.lufthansa.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.lufthansa.com/us/en/flight-search"
            f"?origin={o}&destination={d}&outDate={dep}"
            + (f"&retDate={ret}" if ret else "")
            + f"&pax=1&type={'RT' if ret else 'OW'}"
        ),
    },
    "emirates": {
        "base": "https://www.emirates.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.emirates.com/us/english/book/"
            f"?from={o}&to={d}&departing={dep}"
            + (f"&returning={ret}" if ret else "")
            + f"&pax=1&class=Economy&trip={'return' if ret else 'oneWay'}"
        ),
    },
    "qatar airways": {
        "base": "https://www.qatarairways.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.qatarairways.com/en/booking.html"
            f"?from={o}&to={d}&departing={dep}"
            + (f"&returning={ret}" if ret else "")
            + f"&adults=1&children=0&infants=0"
            + f"&trip={'return' if ret else 'oneway'}&class=Economy"
        ),
    },
    "turkish airlines": {
        "base": "https://www.turkishairlines.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.turkishairlines.com/en-us/flights/"
            f"?portO={o}&portD={d}&departureDateM={dep}"
            + (f"&returnDateM={ret}" if ret else "")
            + f"&adult=1&tripType={'RT' if ret else 'OW'}&cabinClass=ECONOMY"
        ),
    },
    "singapore airlines": {
        "base": "https://www.singaporeair.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.singaporeair.com/en_UK/plan-and-book/booking/"
            f"?from={o}&to={d}&departDate={dep}"
            + (f"&returnDate={ret}" if ret else "")
            + f"&cabinClass=Y&adults=1"
        ),
    },
    "air france": {
        "base": "https://www.airfrance.us",
        "search": lambda o, d, dep, ret: (
            f"https://www.airfrance.us/search/open-dates"
            f"?pax=1:0:0:0:0:0:0:0"
            f"&origin={o}&destination={d}&outboundDate={dep}"
            + (f"&inboundDate={ret}" if ret else "")
        ),
    },
    "klm": {
        "base": "https://www.klm.us",
        "search": lambda o, d, dep, ret: (
            f"https://www.klm.us/search/open-dates"
            f"?pax=1:0:0:0:0:0:0:0"
            f"&origin={o}&destination={d}&outboundDate={dep}"
            + (f"&inboundDate={ret}" if ret else "")
        ),
    },
    "sun country": {
        "base": "https://www.suncountry.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.suncountry.com/book/flights"
        ),
    },
    "breeze": {
        "base": "https://www.flybreeze.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.flybreeze.com/home/booking"
        ),
    },
    "allegiant": {
        "base": "https://www.allegiantair.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.allegiantair.com/flights"
        ),
    },
    "avianca": {
        "base": "https://www.avianca.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.avianca.com/en/booking/"
        ),
    },
    "copa": {
        "base": "https://www.copaair.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.copaair.com/en-us/web/booking"
        ),
    },
    "volaris": {
        "base": "https://www.volaris.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.volaris.com/en/flights"
        ),
    },
    "westjet": {
        "base": "https://www.westjet.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.westjet.com/en-ca/book-trip/flights"
        ),
    },
    "etihad": {
        "base": "https://www.etihad.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.etihad.com/en-us/fly-etihad/book-a-flight"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1&class=Economy&trip={'return' if ret else 'oneWay'}"
        ),
    },
    "korean air": {
        "base": "https://www.koreanair.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.koreanair.com/booking/best-prices"
            f"?tripType={'RT' if ret else 'OW'}&origin={o}&destination={d}"
            f"&departDate={dep}" + (f"&returnDate={ret}" if ret else "")
            + f"&adt=1&cabin=Y"
        ),
    },
    "cathay pacific": {
        "base": "https://www.cathaypacific.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.cathaypacific.com/cx/en_US/book-a-trip/flight-search.html"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1&class=economy"
        ),
    },
    "pia": {
        "base": "https://www.piac.com.pk",
        "search": lambda o, d, dep, ret: (
            f"https://www.piac.com.pk/booking"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1&trip={'RT' if ret else 'OW'}"
        ),
    },
    "air india": {
        "base": "https://www.airindia.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.airindia.com/in/en/book.html"
            f"?from={o}&to={d}&date={dep}"
            + (f"&returnDate={ret}" if ret else "")
            + f"&adults=1&tripType={'RT' if ret else 'OW'}"
        ),
    },
    "saudia": {
        "base": "https://www.saudia.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.saudia.com/booking/flights"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1"
        ),
    },
    "eva air": {
        "base": "https://www.evaair.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.evaair.com/en-us/booking/"
            f"?origin={o}&destination={d}&departureDate={dep}"
            + (f"&returnDate={ret}" if ret else "")
        ),
    },
    "latam": {
        "base": "https://www.latamairlines.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.latamairlines.com/us/en/booking"
            f"?origin={o}&destination={d}&outbound={dep}"
            + (f"&inbound={ret}" if ret else "")
            + f"&adt=1&cabin=Economy"
        ),
    },
    "norwegian": {
        "base": "https://www.norwegian.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.norwegian.com/us/booking/flight-offers/"
            f"?from={o}&to={d}&d_date={dep}"
            + (f"&r_date={ret}" if ret else "")
            + f"&adults=1&tripType={'roundTrip' if ret else 'oneWay'}"
        ),
    },
    "aeromexico": {
        "base": "https://www.aeromexico.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.aeromexico.com/en-us/booking"
            f"?from={o}&to={d}&date={dep}"
            + (f"&returnDate={ret}" if ret else "")
            + f"&adults=1"
        ),
    },
    "iberia": {
        "base": "https://www.iberia.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.iberia.com/us/flights/"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1&cabin=Economy"
        ),
    },
    "aer lingus": {
        "base": "https://www.aerlingus.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.aerlingus.com/booking/select-flights"
            f"?from={o}&to={d}&date={dep}"
            + (f"&returnDate={ret}" if ret else "")
            + f"&adults=1"
        ),
    },
    "finnair": {
        "base": "https://www.finnair.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.finnair.com/en/booking"
            f"?origin={o}&destination={d}&departureDate={dep}"
            + (f"&returnDate={ret}" if ret else "")
            + f"&adults=1"
        ),
    },
    "icelandair": {
        "base": "https://www.icelandair.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.icelandair.com/flights/search/"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1"
        ),
    },
    "thai airways": {
        "base": "https://www.thaiairways.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.thaiairways.com/en/booking/flight-search.page"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1&cabin=Economy"
        ),
    },
    "qantas": {
        "base": "https://www.qantas.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.qantas.com/us/en/book-a-trip/flights.html"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1&cabin=economy"
        ),
    },
    "air new zealand": {
        "base": "https://www.airnewzealand.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.airnewzealand.com/book/flights"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1"
        ),
    },
    "philippine airlines": {
        "base": "https://www.philippineairlines.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.philippineairlines.com/booking"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1"
        ),
    },
    "ethiopian airlines": {
        "base": "https://www.ethiopianairlines.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.ethiopianairlines.com/booking/flight"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1"
        ),
    },
    "royal jordanian": {
        "base": "https://www.rj.com",
        "search": lambda o, d, dep, ret: (
            f"https://www.rj.com/en/book-a-trip/flights"
            f"?from={o}&to={d}&departure={dep}"
            + (f"&return={ret}" if ret else "")
            + f"&adults=1"
        ),
    },
}

# Aliases for fuzzy airline matching
AIRLINE_ALIASES = {
    "delta air lines": "delta",
    "delta airlines": "delta",
    "united airlines": "united",
    "american airlines": "american",
    "southwest airlines": "southwest",
    "jetblue airways": "jetblue",
    "spirit airlines": "spirit",
    "frontier airlines": "frontier",
    "alaska airlines": "alaska",
    "hawaiian airlines": "hawaiian",
    "sun country airlines": "sun country",
    "breeze airways": "breeze",
    "allegiant air": "allegiant",
    "ana": "ana",
    "jal": "jal",
    "japan airlines": "jal",
    "all nippon airways": "ana",
    "pia": "pia",
    "pakistan international": "pia",
    "pakistan international airlines": "pia",
    "serene air": "serene air",
    "airsial": "airsial",
    "tap air portugal": "tap portugal",
    "air arabia": "air arabia",
    "flydubai": "flydubai",
    "indigo": "indigo",
    "spicejet": "spicejet",
    "srilankan airlines": "srilankan",
    "ethiopian airlines": "ethiopian",
    "kenya airways": "kenya airways",
    "royal air maroc": "royal air maroc",
    "egyptair": "egyptair",
    "china eastern": "china eastern",
    "china southern": "china southern",
    "air china": "air china",
    "cathay pacific": "cathay pacific",
    "korean air": "korean air",
    "cebu pacific": "cebu pacific",
    "airasia": "airasia",
    "vietjet": "vietjet",
    "starlux": "starlux",
    "pegasus": "pegasus",
    "wizz air": "wizz air",
    "norse atlantic": "norse atlantic",
}


def _get_airline_booking_url(airline_name: str, origin: str, destination: str,
                              depart_date: str, return_date: str = "") -> str:
    """Get a direct airline booking URL for the given flight details.

    Returns airline-specific deep link if available, otherwise falls back
    to a Google Flights link that filters for that specific airline.
    """
    if not airline_name or airline_name == "—":
        return ""

    key = airline_name.strip().lower()
    # Check aliases first
    if key in AIRLINE_ALIASES:
        key = AIRLINE_ALIASES[key]

    # Direct match
    if key in AIRLINE_BOOKING_URLS:
        try:
            origin_code = _resolve_airport(origin)
            dest_code = _resolve_airport(destination)
            dep = _parse_date(depart_date) if not re.match(r'^\d{4}-\d{2}-\d{2}$', depart_date) else depart_date
            ret = ""
            if return_date:
                ret = _parse_date(return_date) if not re.match(r'^\d{4}-\d{2}-\d{2}$', return_date) else return_date
            return AIRLINE_BOOKING_URLS[key]["search"](origin_code, dest_code, dep, ret)
        except Exception:
            return AIRLINE_BOOKING_URLS[key]["base"]

    # Partial match
    for airline_key in AIRLINE_BOOKING_URLS:
        if airline_key in key or key in airline_key:
            try:
                origin_code = _resolve_airport(origin)
                dest_code = _resolve_airport(destination)
                dep = _parse_date(depart_date) if not re.match(r'^\d{4}-\d{2}-\d{2}$', depart_date) else depart_date
                ret = ""
                if return_date:
                    ret = _parse_date(return_date) if not re.match(r'^\d{4}-\d{2}-\d{2}$', return_date) else return_date
                return AIRLINE_BOOKING_URLS[airline_key]["search"](origin_code, dest_code, dep, ret)
            except Exception:
                return AIRLINE_BOOKING_URLS[airline_key]["base"]

    # Fallback: Google Flights filtered by airline name
    origin_code = _resolve_airport(origin)
    dest_code = _resolve_airport(destination)
    dep = _parse_date(depart_date) if not re.match(r'^\d{4}-\d{2}-\d{2}$', depart_date) else depart_date
    q = f"flights from {origin_code} to {dest_code} on {dep} {airline_name}"
    return f"https://www.google.com/travel/flights?q={urllib.parse.quote_plus(q)}"


def _verify_booking_link(url: str, timeout: int = 5) -> bool:
    """Verify a booking link is valid and the domain exists.

    Light check — just verifies the URL is well-formed and the domain resolves.
    We DON'T do full HTTP checks because airline sites aggressively block bots,
    return 403s, require JS, etc. — but the links still work fine in a browser.
    """
    if not url or "google.com/travel" in url:
        return True  # Google Flights links are always valid

    try:
        import socket
        from urllib.parse import urlparse
        parsed = urlparse(url)
        if not parsed.scheme or not parsed.netloc:
            return False
        # Just verify the domain resolves (DNS check)
        host = parsed.netloc.split(':')[0]
        socket.getaddrinfo(host, None, socket.AF_INET, socket.SOCK_STREAM)
        return True
    except (socket.gaierror, Exception):
        return False


def _get_verified_booking_link(airline_name: str, origin: str, destination: str,
                                depart_date: str, return_date: str = "",
                                search_url: str = "") -> str:
    """Get a booking link and verify it works. Falls back to Google Flights if broken.

    Priority:
      1. Airline-specific deep link (verified)
      2. Google Flights direct search URL
    """
    # Try airline-specific link
    airline_url = _get_airline_booking_url(airline_name, origin, destination, depart_date, return_date)

    if airline_url and "google.com/travel" not in airline_url:
        # Verify the airline link works
        if _verify_booking_link(airline_url):
            return airline_url
        else:
            logger.warning(f"    ⚠️ Broken link for {airline_name}: {airline_url[:60]}... → using Google Flights")

    # Fallback: Google Flights search URL (always works)
    if search_url:
        return search_url

    # Build a Google Flights search URL as final fallback
    origin_code = _resolve_airport(origin)
    dest_code = _resolve_airport(destination)
    dep = _parse_date(depart_date) if not re.match(r'^\d{4}-\d{2}-\d{2}$', depart_date) else depart_date
    return _build_booking_link(origin, destination, dep, return_date)


# ═══════════════════════════════════════════════════════
#  PHASE 12 — Structured DOM Flight Parser
# ═══════════════════════════════════════════════════════

# JavaScript injected into Google Flights via CDP to extract structured JSON
# directly from the DOM instead of regex-parsing raw innerText.
_DOM_EXTRACT_JS = r"""
(function() {
    var flights = [];
    var seen = {};

    // Google Flights price insight banner (e.g. "Prices are currently low")
    var priceInsight = '';
    var insightEls = document.querySelectorAll('[class*="price-insight"], [class*="advisory"], [data-lk]');
    insightEls.forEach(function(el) {
        var t = el.innerText.trim().toLowerCase();
        if ((t.includes('prices are') || t.includes('price is') || t.includes('typically cost') || t.includes('is low') || t.includes('is high') || t.includes('cheaper than usual') || t.includes('more expensive')) && t.length < 300) {
            priceInsight = el.innerText.trim();
        }
    });
    // Also check for the insights section
    if (!priceInsight) {
        var allText = document.body.innerText;
        var insightMatch = allText.match(/(Prices are (?:currently |typically )?(?:low|high|typical|higher than usual|lower than usual)[^.]*\.)/i);
        if (insightMatch) priceInsight = insightMatch[1];
        if (!priceInsight) {
            var costMatch = allText.match(/((?:Flights |These flights |Prices )(?:typically cost|usually cost|are priced)[^.]*\.)/i);
            if (costMatch) priceInsight = costMatch[1];
        }
    }

    // Try data-resultid elements first (most reliable)
    var cards = document.querySelectorAll('[data-resultid]');
    if (cards.length === 0) {
        cards = document.querySelectorAll('li[class*="pIav2d"], ul[class*="Rk10dc"] > li, [role="listitem"]');
    }
    if (cards.length === 0) {
        cards = document.querySelectorAll('li');
    }

    cards.forEach(function(card) {
        var text = card.innerText.trim();
        if (!text.includes('$') || text.length < 20 || text.length > 800) return;

        var flight = {};

        // Price — look for $ followed by digits
        var priceMatch = text.match(/\$[\d,]+/);
        if (priceMatch) flight.price = priceMatch[0];
        else return;

        // Times — "HH:MM AM/PM"
        var times = text.match(/\d{1,2}:\d{2}\s*(?:AM|PM)/gi);
        if (times && times.length >= 2) {
            flight.depart_time = times[0];
            flight.arrive_time = times[1];
        }

        // Duration — "X hr Y min" or "Xhr Ymin"
        var durMatch = text.match(/(\d+)\s*h(?:r|rs?)?\s*(?:(\d+)\s*m(?:in)?)?/i);
        if (durMatch) {
            flight.duration = durMatch[0].trim();
        }

        // Stops — "Nonstop" or "N stop(s)"
        if (/\bnonstop\b/i.test(text)) {
            flight.stops = 'Nonstop';
        } else {
            var stopMatch = text.match(/(\d+)\s*stop/i);
            if (stopMatch) {
                var n = parseInt(stopMatch[1]);
                flight.stops = n + ' stop' + (n > 1 ? 's' : '');
            }
        }

        // Layover details — "Stop in XXX" or "Layover X hr in XXX"
        var layoverMatch = text.match(/(?:stop|layover|connection)\s+(?:(\d+\s*h(?:r|rs?)?\s*(?:\d+\s*m(?:in)?)?)\s+)?(?:in|at)\s+([A-Z]{3}(?:\s*,\s*[A-Z]{3})*)/i);
        if (layoverMatch) {
            flight.layover_duration = layoverMatch[1] || '';
            flight.layover_airport = layoverMatch[2] || '';
        }
        // Also try: "2 hr 15 min DEN" pattern (layover duration + airport code)
        if (!flight.layover_airport && flight.stops && flight.stops !== 'Nonstop') {
            var layPattern = text.match(/(\d+\s*h(?:r|rs?)?\s*\d*\s*m?(?:in)?)\s+([A-Z]{3})\b/);
            if (layPattern && layPattern[2] !== flight.depart_time) {
                flight.layover_duration = layPattern[1];
                flight.layover_airport = layPattern[2];
            }
        }

        // Fare class / cabin
        var fareTexts = ['basic economy', 'main cabin', 'economy', 'premium economy',
                        'business', 'first class', 'basic', 'comfort+', 'delta one',
                        'polaris', 'mint', 'flagship'];
        var textLower = text.toLowerCase();
        for (var fi = 0; fi < fareTexts.length; fi++) {
            if (textLower.includes(fareTexts[fi])) {
                flight.fare_class = fareTexts[fi].charAt(0).toUpperCase() + fareTexts[fi].slice(1);
                break;
            }
        }

        // Baggage — check for carry-on / checked bag mentions
        if (/no carry-on/i.test(text)) {
            flight.baggage = 'No carry-on';
        } else if (/carry-on bag/i.test(text) && /no checked bag/i.test(text)) {
            flight.baggage = 'Carry-on only';
        } else if (/checked bag/i.test(text)) {
            flight.baggage = 'Checked bag included';
        } else if (/carry-on/i.test(text)) {
            flight.baggage = 'Carry-on included';
        }

        // Airline — check known airlines list
        var airlines = [
            'Delta', 'United', 'American', 'Southwest', 'JetBlue', 'Spirit',
            'Frontier', 'Alaska', 'Hawaiian', 'Sun Country', 'Breeze', 'Allegiant',
            'Air Canada', 'WestJet',
            'British Airways', 'Lufthansa', 'Air France', 'KLM', 'Iberia',
            'Ryanair', 'EasyJet', 'Norwegian', 'SAS', 'Swiss', 'Austrian',
            'TAP Portugal', 'TAP Air Portugal', 'Aer Lingus', 'Finnair', 'Icelandair', 'Condor',
            'Emirates', 'Qatar Airways', 'Turkish Airlines', 'Etihad', 'Saudia',
            'Royal Jordanian', 'Oman Air', 'Gulf Air', 'flynas', 'Air Arabia',
            'Singapore Airlines', 'ANA', 'JAL', 'Japan Airlines',
            'All Nippon Airways', 'Cathay Pacific', 'Korean Air',
            'Asiana', 'China Airlines', 'EVA Air', 'Thai Airways',
            'Vietnam Airlines', 'Philippine Airlines', 'Malaysia Airlines',
            'Garuda Indonesia', 'Air India', 'IndiGo', 'SpiceJet',
            'PIA', 'Pakistan International', 'Serene Air', 'AirSial', 'airblue',
            'Avianca', 'Copa', 'Volaris', 'VivaAerobus', 'LATAM',
            'Aeromexico', 'GOL', 'Azul',
            'Qantas', 'Air New Zealand', 'Fiji Airways',
            'Pegasus', 'Norse Atlantic', 'Play', 'Transavia',
            'Vueling', 'Wizz Air', 'LOT Polish', 'Czech Airlines',
            'flydubai', 'Jazeera Airways', 'SriLankan Airlines',
            'Nepal Airlines', 'Biman Bangladesh', 'Air China',
            'China Eastern', 'China Southern', 'Hainan Airlines',
            'Xiamen Airlines', 'Sichuan Airlines', 'Spring Airlines',
            'Bamboo Airways', 'Cebu Pacific', 'AirAsia', 'Scoot',
            'Jetstar', 'Tiger Air', 'Peach Aviation',
            'Ethiopian Airlines', 'Kenya Airways', 'South African Airways',
            'Royal Air Maroc', 'EgyptAir', 'Tunisair',
            'Caribbean Airlines', 'JetSMART', 'SKY Airline',
            'Starlux', 'Vietjet'
        ];
        var lines = text.split('\n');
        for (var li = 0; li < lines.length && li < 15; li++) {
            var lineTrim = lines[li].trim();
            for (var ai = 0; ai < airlines.length; ai++) {
                if (lineTrim.toLowerCase() === airlines[ai].toLowerCase() ||
                    (lineTrim.toLowerCase().includes(airlines[ai].toLowerCase()) && lineTrim.length < 40)) {
                    flight.airline = airlines[ai];
                    break;
                }
            }
            if (flight.airline) break;
        }

        // Codeshare / operated by
        var operatedMatch = text.match(/(?:operated|marketed)\s+by\s+([A-Za-z\s]+?)(?:\.|,|\n)/i);
        if (operatedMatch) {
            flight.operated_by = operatedMatch[1].trim();
        }

        // Airport codes from text
        var airportCodes = text.match(/\b([A-Z]{3})\b/g);
        if (airportCodes) {
            var skip = ['THE','AND','FOR','NOT','ALL','NEW','USD','AVG','TOP','SEE',
                       'MON','TUE','WED','THU','FRI','SAT','SUN','JAN','FEB','MAR',
                       'APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC','EST','PST','CST','MST'];
            var codes = airportCodes.filter(function(c) { return skip.indexOf(c) === -1; });
            if (codes.length >= 2) {
                flight.from_airport = codes[0];
                flight.to_airport = codes[codes.length > 2 ? codes.length - 1 : 1];
            }
        }

        // Emissions
        var emissionMatch = text.match(/(\d+)\s*kg\s*CO/i);
        if (emissionMatch) {
            flight.emissions_kg = parseInt(emissionMatch[1]);
        }

        // Dedup key
        var key = (flight.price || '') + '_' + (flight.airline || '') + '_' + (flight.depart_time || '');
        if (seen[key]) return;
        seen[key] = true;

        if (flight.price) flights.push(flight);
    });

    // Sort by price
    flights.sort(function(a, b) {
        var pa = parseInt((a.price || '$99999').replace(/[^0-9]/g, ''));
        var pb = parseInt((b.price || '$99999').replace(/[^0-9]/g, ''));
        return pa - pb;
    });

    return JSON.stringify({flights: flights, priceInsight: priceInsight, count: flights.length});
})()
"""

# JavaScript for extracting return flight info (second leg of round-trip)
_RETURN_FLIGHT_JS = r"""
(function() {
    // Google Flights often shows "Departing" and "Returning" sections
    var text = document.body.innerText;
    var returnInfo = {};

    // Look for return flight section
    var returnMatch = text.match(/(?:Return|Returning|Return flight)[:\s]*\n?([\s\S]{0,500})/i);
    if (returnMatch) {
        var rText = returnMatch[1];
        var rPrice = rText.match(/\$[\d,]+/);
        var rTimes = rText.match(/\d{1,2}:\d{2}\s*(?:AM|PM)/gi);
        var rDur = rText.match(/(\d+)\s*h(?:r|rs?)?\s*(?:(\d+)\s*m(?:in)?)?/i);
        var rStop = rText.match(/(\d+)\s*stop/i);
        var rNonstop = /\bnonstop\b/i.test(rText);

        if (rTimes && rTimes.length >= 2) {
            returnInfo.depart_time = rTimes[0];
            returnInfo.arrive_time = rTimes[1];
        }
        if (rDur) returnInfo.duration = rDur[0].trim();
        if (rNonstop) returnInfo.stops = 'Nonstop';
        else if (rStop) returnInfo.stops = rStop[1] + ' stop' + (parseInt(rStop[1]) > 1 ? 's' : '');
    }

    return JSON.stringify(returnInfo);
})()
"""


def _extract_flight_data_dom(cdp) -> dict:
    """Extract structured flight data from Google Flights via DOM parsing.

    Returns: {"flights": [...], "price_insight": str, "return_flight": dict}
    Much more reliable than regex on raw page text.
    """
    result = {"flights": [], "price_insight": "", "return_flight": {}}

    try:
        r = cdp.send("Runtime.evaluate", {
            "expression": _DOM_EXTRACT_JS,
            "returnByValue": True,
        })
        raw = r.get("result", {}).get("value", "{}")
        data = json.loads(raw) if isinstance(raw, str) else raw
        result["flights"] = data.get("flights", [])
        result["price_insight"] = data.get("priceInsight", "")
    except Exception as e:
        logger.warning(f"    ⚠️ DOM extraction failed: {e}")

    # Try to get return flight info for round-trips
    try:
        r2 = cdp.send("Runtime.evaluate", {
            "expression": _RETURN_FLIGHT_JS,
            "returnByValue": True,
        })
        raw2 = r2.get("result", {}).get("value", "{}")
        ret_data = json.loads(raw2) if isinstance(raw2, str) else raw2
        if ret_data and any(ret_data.values()):
            result["return_flight"] = ret_data
    except Exception:
        pass

    return result


def _extract_flight_data(page_text: str) -> list:
    """Fallback parser: regex on raw page text (used when DOM extraction fails)."""
    flights = []
    lines = page_text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        price_match = re.search(r'\$[\d,]+', line)
        if price_match:
            price = price_match.group()
            context_start = max(0, i - 15)
            context_end = min(len(lines), i + 3)
            context = "\n".join(lines[context_start:context_end])
            time_matches = re.findall(r'\d{1,2}:\d{2}\s*(?:AM|PM)', context)
            airport_matches = re.findall(r'\b([A-Z]{3})\b', context)
            airport_codes = [a for a in airport_matches if a not in (
                'THE', 'AND', 'FOR', 'NOT', 'ALL', 'NEW', 'USD', 'AVG',
                'TOP', 'SEE', 'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN',
            )]
            duration_match = re.search(r'(\d+\s*hr?\s*(?:\d+\s*min)?)', context)
            duration = duration_match.group(1).strip() if duration_match else ""
            airline = ""
            common_airlines = [
                # US Carriers
                "Delta", "United", "American", "Southwest", "JetBlue", "Spirit",
                "Frontier", "Alaska", "Hawaiian", "Sun Country", "Breeze", "Allegiant",
                # Canadian
                "Air Canada", "WestJet",
                # European
                "British Airways", "Lufthansa", "Air France", "KLM", "Iberia",
                "Ryanair", "EasyJet", "Norwegian", "SAS", "Swiss", "Austrian",
                "TAP Portugal", "TAP Air Portugal", "Aer Lingus", "Finnair", "Icelandair", "Condor",
                "Vueling", "Wizz Air", "LOT Polish", "Transavia", "Play",
                # Middle East
                "Emirates", "Qatar Airways", "Turkish Airlines", "Etihad", "Saudia",
                "Royal Jordanian", "Oman Air", "Gulf Air", "flynas", "Air Arabia",
                "flydubai", "Jazeera Airways",
                # Asian
                "Singapore Airlines", "ANA", "JAL", "Japan Airlines",
                "All Nippon Airways", "Cathay Pacific", "Korean Air",
                "Asiana", "China Airlines", "EVA Air", "Thai Airways",
                "Vietnam Airlines", "Philippine Airlines", "Malaysia Airlines",
                "Garuda Indonesia", "Air India", "IndiGo", "SpiceJet",
                "Air China", "China Eastern", "China Southern", "Hainan Airlines",
                "Starlux", "Vietjet", "Bamboo Airways", "Cebu Pacific",
                "AirAsia", "Scoot", "Jetstar", "Peach Aviation",
                # South Asia
                "PIA", "Pakistan International", "Serene Air", "AirSial", "airblue",
                "SriLankan Airlines", "Nepal Airlines", "Biman Bangladesh",
                # Latin American
                "Avianca", "Copa", "Volaris", "VivaAerobus", "LATAM",
                "Aeromexico", "GOL", "Azul", "JetSMART", "SKY Airline",
                # Oceania
                "Qantas", "Air New Zealand", "Fiji Airways",
                # African
                "Ethiopian Airlines", "Kenya Airways", "South African Airways",
                "Royal Air Maroc", "EgyptAir",
                # Other
                "Norse Atlantic", "Pegasus",
            ]
            for li in range(max(0, i - 12), i):
                line_text = lines[li].strip()
                for al in common_airlines:
                    if al.lower() == line_text.lower() or al.lower() in line_text.lower():
                        airline = al
                        break
                if airline:
                    break
            stops_text = "Nonstop"
            stop_match = re.search(r'(\d+)\s*stop', context, re.IGNORECASE)
            if stop_match:
                n = stop_match.group(1)
                stops_text = f"{n} stop{'s' if int(n) > 1 else ''}"
            elif "nonstop" in context.lower():
                stops_text = "Nonstop"
            flight = {
                "price": price,
                "airline": airline or "—",
                "stops": stops_text,
                "duration": duration or "—",
            }
            if len(time_matches) >= 2:
                flight["depart_time"] = time_matches[0]
                flight["arrive_time"] = time_matches[1]
            if len(airport_codes) >= 2:
                flight["from"] = airport_codes[0]
                flight["to"] = airport_codes[1]
            key = f"{price}_{airline}_{flight.get('depart_time', '')}"
            if not any(f"{f['price']}_{f['airline']}_{f.get('depart_time', '')}" == key for f in flights):
                flights.append(flight)
        i += 1

    def price_num(f):
        try:
            return int(f["price"].replace("$", "").replace(",", ""))
        except (ValueError, KeyError):
            return 99999
    flights.sort(key=price_num)
    return flights


def _price_num(price_str: str) -> int:
    """Extract numeric price from string like '$542'."""
    try:
        return int(re.sub(r'[^\d]', '', price_str))
    except (ValueError, TypeError):
        return 99999


def _format_flights(flights: list, query_desc: str) -> str:
    """Format flights into a clean, readable report."""
    if not flights:
        return f"❌ No flights found for: {query_desc}\n\nTry different dates or a nearby airport."
    lines = [f"✈️ **Flight Search Results** — {query_desc}\n"]
    lines.append(f"Found {len(flights)} options (sorted by price):\n")
    lines.append(f"{'#':<3} {'Price':<10} {'Airline':<15} {'Depart':<10} {'Arrive':<10} {'Duration':<12} {'Stops'}")
    lines.append("─" * 80)
    for i, f in enumerate(flights[:15], 1):
        lines.append(
            f"{i:<3} {f.get('price', '?'):<10} "
            f"{f.get('airline', '?'):<15} "
            f"{f.get('depart_time', '?'):<10} "
            f"{f.get('arrive_time', '?'):<10} "
            f"{f.get('duration', '?'):<12} "
            f"{f.get('stops', '?')}"
        )
    if len(flights) > 15:
        lines.append(f"\n... and {len(flights) - 15} more options")
    cheapest = flights[0]
    lines.append(f"\n💰 **Cheapest**: {cheapest.get('price', '?')} — {cheapest.get('airline', '?')} ({cheapest.get('stops', '?')}, {cheapest.get('duration', '?')})")
    nonstops = [f for f in flights if "nonstop" in f.get("stops", "").lower()]
    if nonstops:
        best_ns = nonstops[0]
        lines.append(f"✈️ **Cheapest Nonstop**: {best_ns.get('price', '?')} — {best_ns.get('airline', '?')} ({best_ns.get('duration', '?')})")
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════
#  PHASE 11 — Intelligence Layer: Suggestions & Analytics
# ═══════════════════════════════════════════════════════

# Nearby airport alternatives — major US + international hubs
NEARBY_AIRPORTS = {
    "JFK": ["EWR", "LGA"], "EWR": ["JFK", "LGA"], "LGA": ["JFK", "EWR"],
    "LAX": ["SNA", "BUR", "LGB", "ONT"], "SNA": ["LAX", "LGB"],
    "ORD": ["MDW"], "MDW": ["ORD"],
    "SFO": ["OAK", "SJC"], "OAK": ["SFO", "SJC"], "SJC": ["SFO", "OAK"],
    "MIA": ["FLL", "PBI"], "FLL": ["MIA", "PBI"],
    "DFW": ["DAL", "AUS", "SAT"], "DAL": ["DFW"],
    "IAH": ["HOU"], "HOU": ["IAH"],
    "DCA": ["IAD", "BWI"], "IAD": ["DCA", "BWI"], "BWI": ["DCA", "IAD"],
    "TPA": ["MCO", "SRQ", "PIE"], "MCO": ["TPA", "SFB"],
    "SEA": ["PAE"], "BOS": ["PVD", "MHT"],
    "ATL": [], "DEN": ["COS"],
    "PHX": ["AZA", "TUS"], "DTW": ["FNT"],
    "MSP": [],
    "LHR": ["LGW", "STN", "LTN"], "LGW": ["LHR", "STN"],
    "CDG": ["ORY"], "ORY": ["CDG"],
    "NRT": ["HND"], "HND": ["NRT"],
    "ICN": ["GMP"], "PEK": ["PKX"],
}

AIRPORT_NAMES = {
    "JFK": "JFK", "EWR": "Newark", "LGA": "LaGuardia",
    "LAX": "LAX", "SNA": "Orange County", "BUR": "Burbank", "LGB": "Long Beach", "ONT": "Ontario",
    "ORD": "O'Hare", "MDW": "Midway",
    "SFO": "SFO", "OAK": "Oakland", "SJC": "San Jose",
    "MIA": "Miami", "FLL": "Fort Lauderdale", "PBI": "West Palm Beach",
    "DFW": "DFW", "DAL": "Love Field",
    "IAH": "IAH", "HOU": "Hobby",
    "DCA": "Reagan", "IAD": "Dulles", "BWI": "Baltimore",
    "TPA": "Tampa", "MCO": "Orlando", "SRQ": "Sarasota", "PIE": "St. Pete",
    "SEA": "Seattle", "BOS": "Boston", "PVD": "Providence", "MHT": "Manchester",
    "LHR": "Heathrow", "LGW": "Gatwick", "STN": "Stansted", "LTN": "Luton",
    "CDG": "CDG", "ORY": "Orly",
    "NRT": "Narita", "HND": "Haneda",
    "ICN": "Incheon", "GMP": "Gimpo",
    "SFB": "Sanford", "PAE": "Paine Field",
    "COS": "Colorado Springs", "AZA": "Mesa-Gateway", "TUS": "Tucson",
    "FNT": "Flint", "PKX": "Daxing",
}


def _analyze_flights(flights: list, origin_code: str, dest_code: str,
                      depart_date: str, return_date: str = "") -> dict:
    """Deep analytics on flight results — prices, airlines, value scores, insights."""
    if not flights:
        return {"suggestions": [], "analytics": {}}

    prices = [_price_num(f.get("price", "$99999")) for f in flights if _price_num(f.get("price", "$99999")) < 99999]
    if not prices:
        return {"suggestions": [], "analytics": {}}

    avg_price = sum(prices) / len(prices)
    min_price = min(prices)
    max_price = max(prices)
    median_price = sorted(prices)[len(prices) // 2]
    price_range = max_price - min_price
    std_dev = math.sqrt(sum((p - avg_price) ** 2 for p in prices) / len(prices)) if len(prices) > 1 else 0

    # Airline breakdown
    airline_stats = {}
    for f in flights:
        al = f.get("airline", "—")
        p = _price_num(f.get("price", "$99999"))
        if p >= 99999:
            continue
        if al not in airline_stats:
            airline_stats[al] = {"prices": [], "nonstop": 0, "stops": 0}
        airline_stats[al]["prices"].append(p)
        if "nonstop" in f.get("stops", "").lower():
            airline_stats[al]["nonstop"] += 1
        else:
            airline_stats[al]["stops"] += 1

    for al in airline_stats:
        s = airline_stats[al]
        s["avg"] = sum(s["prices"]) / len(s["prices"])
        s["min"] = min(s["prices"])
        s["count"] = len(s["prices"])

    # Nonstop analysis
    nonstops = [f for f in flights if "nonstop" in f.get("stops", "").lower()]
    nonstop_prices = [_price_num(f.get("price", "$99999")) for f in nonstops if _price_num(f.get("price", "$99999")) < 99999]
    connecting = [f for f in flights if "nonstop" not in f.get("stops", "").lower()]
    connecting_prices = [_price_num(f.get("price", "$99999")) for f in connecting if _price_num(f.get("price", "$99999")) < 99999]

    nonstop_premium = 0
    if nonstop_prices and connecting_prices:
        nonstop_premium = min(nonstop_prices) - min(connecting_prices)

    # Value scores — price vs convenience trade-off
    for f in flights:
        p = _price_num(f.get("price", "$99999"))
        if p >= 99999:
            f["value_score"] = 0
            f["value_label"] = "—"
            continue

        # Score: 0-100. Lower price = higher score, nonstop = bonus, short duration = bonus
        price_score = max(0, 100 - ((p - min_price) / max(price_range, 1)) * 70)

        stop_bonus = 15 if "nonstop" in f.get("stops", "").lower() else 0
        dur_text = f.get("duration", "")
        dur_mins = 0
        hr_m = re.search(r'(\d+)\s*hr?', dur_text)
        min_m = re.search(r'(\d+)\s*min', dur_text)
        if hr_m:
            dur_mins = int(hr_m.group(1)) * 60
        if min_m:
            dur_mins += int(min_m.group(1))
        dur_bonus = max(0, 15 - (dur_mins / 60)) if dur_mins > 0 else 0

        # v5.0 — Layover quality bonus (short layover = better for connecting)
        layover_bonus = 0
        if f.get("layover_duration") and "nonstop" not in f.get("stops", "").lower():
            lay_text = f.get("layover_duration", "")
            lay_mins = 0
            lay_hr = re.search(r'(\d+)\s*hr?', lay_text)
            lay_min = re.search(r'(\d+)\s*min', lay_text)
            if lay_hr:
                lay_mins = int(lay_hr.group(1)) * 60
            if lay_min:
                lay_mins += int(lay_min.group(1))
            if 60 <= lay_mins <= 150:
                layover_bonus = 5  # Ideal layover (1-2.5h)
            elif lay_mins > 300:
                layover_bonus = -5  # Long layover penalty

        # v5.0 — Baggage bonus
        baggage_bonus = 0
        bag = f.get("baggage", "").lower()
        if "checked" in bag:
            baggage_bonus = 3
        elif "carry-on" in bag and "no" not in bag:
            baggage_bonus = 1

        score = min(100, max(0, price_score + stop_bonus + dur_bonus + layover_bonus + baggage_bonus))
        f["value_score"] = round(score)
        if score >= 85:
            f["value_label"] = "🟢 Excellent"
        elif score >= 70:
            f["value_label"] = "🔵 Great"
        elif score >= 50:
            f["value_label"] = "🟡 Good"
        elif score >= 30:
            f["value_label"] = "🟠 Fair"
        else:
            f["value_label"] = "🔴 Pricey"

    # Build suggestions
    suggestions = []

    # 1. Nonstop premium analysis
    if nonstop_prices and connecting_prices and nonstop_premium > 0:
        if nonstop_premium > 100:
            suggestions.append({
                "type": "nonstop_premium", "icon": "💡",
                "text": f"Nonstop flights cost ${nonstop_premium} more than connecting. A 1-stop flight could save you significantly.",
                "priority": 2,
            })
        elif nonstop_premium < 30:
            suggestions.append({
                "type": "nonstop_value", "icon": "✈️",
                "text": f"Nonstop is only ${nonstop_premium} more — worth the time savings!",
                "priority": 1,
            })

    # 2. Price spread — if big range, flag that timing matters
    if price_range > 200:
        savings_pct = round((price_range / max_price) * 100)
        suggestions.append({
            "type": "price_spread", "icon": "📊",
            "text": f"Price range is ${price_range} (${min_price}–${max_price}). Picking wisely saves up to {savings_pct}%.",
            "priority": 1,
        })

    # 3. Nearby airport alternatives
    alt_airports = NEARBY_AIRPORTS.get(origin_code, [])
    if alt_airports:
        alt_names = [f"{AIRPORT_NAMES.get(a, a)} ({a})" for a in alt_airports[:3]]
        suggestions.append({
            "type": "nearby_airports", "icon": "🗺️",
            "text": f"Also check flights from nearby airports: {', '.join(alt_names)}. Prices often differ by $50–$200+.",
            "airports": alt_airports[:3],
            "priority": 3,
        })
    alt_dest = NEARBY_AIRPORTS.get(dest_code, [])
    if alt_dest:
        alt_names = [f"{AIRPORT_NAMES.get(a, a)} ({a})" for a in alt_dest[:3]]
        suggestions.append({
            "type": "nearby_dest", "icon": "📍",
            "text": f"Consider flying into {', '.join(alt_names)} instead — could be cheaper with a short transfer.",
            "airports": alt_dest[:3],
            "priority": 3,
        })

    # 4. Day-of-week insight
    try:
        dep = datetime.strptime(depart_date, "%Y-%m-%d")
        day_name = dep.strftime("%A")
        if dep.weekday() in (4, 5, 6):  # Fri, Sat, Sun
            suggestions.append({
                "type": "day_shift", "icon": "📅",
                "text": f"You're flying on a {day_name}. Weekday departures (Tue–Thu) are typically 10–30% cheaper.",
                "priority": 2,
            })
        elif dep.weekday() in (1, 2, 3):  # Tue, Wed, Thu
            suggestions.append({
                "type": "good_day", "icon": "👍",
                "text": f"Flying on {day_name} — smart! Midweek flights tend to have the best prices.",
                "priority": 4,
            })
    except (ValueError, TypeError):
        pass

    # 5. Advance booking insight
    try:
        dep = datetime.strptime(depart_date, "%Y-%m-%d")
        days_out = (dep - datetime.now()).days
        if days_out < 7:
            suggestions.append({
                "type": "last_minute", "icon": "⚠️",
                "text": "Last-minute booking! Prices are likely inflated. If flexible, booking 2–3 weeks out typically saves 20–40%.",
                "priority": 1,
            })
        elif days_out < 14:
            suggestions.append({
                "type": "soon", "icon": "⏰",
                "text": f"Flying in {days_out} days. Prices may rise further — book soon if the price looks good.",
                "priority": 2,
            })
        elif 21 <= days_out <= 90:
            suggestions.append({
                "type": "sweet_spot", "icon": "🎯",
                "text": f"Booking {days_out} days out — this is the sweet spot for domestic flights. Good timing!",
                "priority": 4,
            })
        elif days_out > 180:
            suggestions.append({
                "type": "early_bird", "icon": "🐦",
                "text": f"Booking {days_out} days ahead. Airlines haven't optimized prices yet — set a tracker for even better deals.",
                "priority": 3,
            })
    except (ValueError, TypeError):
        pass

    # 6. Best overall value pick
    scored = sorted([f for f in flights if f.get("value_score", 0) > 0],
                     key=lambda x: x["value_score"], reverse=True)
    if scored and scored[0] != flights[0]:
        best_val = scored[0]
        suggestions.append({
            "type": "best_value", "icon": "⭐",
            "text": f"Best overall value: {best_val.get('airline', '?')} at {best_val.get('price', '?')} "
                    f"({best_val.get('stops', '?')}, {best_val.get('duration', '?')}) — "
                    f"score {best_val['value_score']}/100.",
            "priority": 1,
        })

    # 7. Multi-airline insight
    if len(airline_stats) >= 3:
        cheapest_airline = min(airline_stats.items(), key=lambda x: x[1]["min"])
        suggestions.append({
            "type": "airline_compare", "icon": "🏢",
            "text": f"{len(airline_stats)} airlines on this route. {cheapest_airline[0]} has the lowest starting at ${cheapest_airline[1]['min']}.",
            "priority": 3,
        })

    # Sort by priority (lower = more important)
    suggestions.sort(key=lambda s: s.get("priority", 5))

    analytics = {
        "price_min": min_price, "price_max": max_price,
        "price_avg": round(avg_price), "price_median": median_price,
        "price_range": price_range, "price_std_dev": round(std_dev),
        "total_flights": len(flights),
        "nonstop_count": len(nonstops), "connecting_count": len(connecting),
        "nonstop_cheapest": min(nonstop_prices) if nonstop_prices else None,
        "connecting_cheapest": min(connecting_prices) if connecting_prices else None,
        "nonstop_premium": nonstop_premium if nonstop_prices and connecting_prices else None,
        "airline_count": len(airline_stats),
        "airline_stats": {k: {"min": v["min"], "avg": round(v["avg"]), "count": v["count"],
                              "nonstop": v["nonstop"]} for k, v in airline_stats.items()},
    }

    return {"suggestions": suggestions, "analytics": analytics}


def _format_flights_rich(flights: list, query_desc: str, origin_code: str,
                          dest_code: str, depart_date: str, return_date: str = "",
                          price_insight: str = "", return_flight: dict = None,
                          tracker_suggestion: str = "") -> str:
    """Format flights into a rich report with analytics, suggestions, and v5.0 details."""
    if not flights:
        return f"❌ No flights found for: {query_desc}\n\nTry different dates or a nearby airport."

    intel = _analyze_flights(flights, origin_code, dest_code, depart_date, return_date)
    analytics = intel["analytics"]
    suggestions = intel["suggestions"]
    cheapest = flights[0]
    nonstops = [f for f in flights if "nonstop" in f.get("stops", "").lower()]

    lines = [f"✈️ **Flight Search Results** — {query_desc}\n"]

    # v5.0 — Google Price Insight banner
    if price_insight:
        lines.append(f"📈 **Google Insight**: {price_insight}\n")

    # Quick stats bar
    lines.append(f"📊 {analytics['total_flights']} flights · "
                 f"${analytics['price_min']}–${analytics['price_max']} · "
                 f"avg ${analytics['price_avg']} · "
                 f"{analytics['nonstop_count']} nonstop · "
                 f"{analytics['airline_count']} airlines\n")

    # Flight table with value scores + new v5.0 columns
    lines.append(f"{'#':<3} {'Price':<10} {'Airline':<15} {'Depart':<10} {'Arrive':<10} {'Duration':<12} {'Stops':<12} {'Layover':<14} {'Value'}")
    lines.append("─" * 100)
    for i, f in enumerate(flights[:15], 1):
        layover = ""
        if f.get("layover_airport"):
            layover = f"via {f['layover_airport']}"
            if f.get("layover_duration"):
                layover += f" {f['layover_duration']}"
        elif "nonstop" in f.get("stops", "").lower():
            layover = "—"
        fare_bag = ""
        if f.get("fare_class"):
            fare_bag = f.get("fare_class", "")
        if f.get("baggage"):
            fare_bag += f" [{f['baggage']}]" if fare_bag else f.get("baggage", "")

        lines.append(
            f"{i:<3} {f.get('price', '?'):<10} "
            f"{f.get('airline', '?'):<15} "
            f"{f.get('depart_time', '?'):<10} "
            f"{f.get('arrive_time', '?'):<10} "
            f"{f.get('duration', '?'):<12} "
            f"{f.get('stops', '?'):<12} "
            f"{layover:<14} "
            f"{f.get('value_label', '')}"
        )
        if fare_bag:
            lines.append(f"{'':>3} {'':>10} {'':>15} 🎫 {fare_bag}")
    if len(flights) > 15:
        lines.append(f"\n... and {len(flights) - 15} more options")

    # Key picks
    lines.append(f"\n💰 **Cheapest**: {cheapest.get('price', '?')} — {cheapest.get('airline', '?')} ({cheapest.get('stops', '?')}, {cheapest.get('duration', '?')})")
    if nonstops:
        best_ns = nonstops[0]
        lines.append(f"✈️ **Cheapest Nonstop**: {best_ns.get('price', '?')} — {best_ns.get('airline', '?')} ({best_ns.get('duration', '?')})")

    scored = sorted([f for f in flights if f.get("value_score", 0) > 0],
                     key=lambda x: x["value_score"], reverse=True)
    if scored:
        bv = scored[0]
        lines.append(f"⭐ **Best Value**: {bv.get('price', '?')} — {bv.get('airline', '?')} ({bv.get('stops', '?')}, {bv.get('duration', '?')}) — Score: {bv['value_score']}/100")

    # v5.0 — Return flight info
    if return_flight and return_flight.get("depart_time"):
        rf = return_flight
        lines.append(f"\n🔄 **Return Flight**: {rf.get('depart_time', '?')} → {rf.get('arrive_time', '?')} · "
                      f"{rf.get('duration', '?')} · {rf.get('stops', '?')}")

    # v5.0 — Tracker suggestion
    if tracker_suggestion:
        lines.append(f"\n{tracker_suggestion}")

    # Suggestions
    if suggestions:
        lines.append("\n💡 **Smart Suggestions:**")
        for s in suggestions[:5]:
            lines.append(f"  {s['icon']} {s['text']}")

    return "\n".join(lines)

def _html_flight_report_email(origin_code, dest_code, depart_date, return_date, flights, search_url,
                               price_insight="", return_flight=None, tracker_suggestion=""):
    """Generate a premium HTML email with analytics, value scores, price chart, suggestions, and v5.0 intel."""
    if not flights:
        return "<p>No flights found.</p>"

    intel = _analyze_flights(flights, origin_code, dest_code, depart_date, return_date)
    analytics = intel["analytics"]
    suggestions = intel["suggestions"]

    cheapest = flights[0]
    nonstops = [f for f in flights if "nonstop" in f.get("stops", "").lower()]
    best_nonstop = nonstops[0] if nonstops else None
    date_display = depart_date
    if return_date:
        date_display = f"{depart_date}  →  {return_date}"

    # Build inline price bar chart (top 8 flights)
    price_bars = ""
    chart_flights = flights[:8]
    chart_max = max(_price_num(f.get("price", "$1")) for f in chart_flights) if chart_flights else 1
    for i, f in enumerate(chart_flights):
        p = _price_num(f.get("price", "$0"))
        pct = max(15, round((p / max(chart_max, 1)) * 100))
        bar_color = "#059669" if i == 0 else "#2563EB" if "nonstop" in f.get("stops", "").lower() else "#64748B"
        label = f.get("airline", "?")[:12]
        price_bars += f"""
        <tr>
          <td style="padding:4px 8px 4px 0;color:#475569;font-size:11px;font-weight:500;white-space:nowrap;width:80px;">{label}</td>
          <td style="padding:4px 0;width:100%;">
            <div style="background:{bar_color};height:22px;width:{pct}%;border-radius:4px;display:flex;align-items:center;justify-content:flex-end;padding-right:8px;">
              <span style="color:#fff;font-size:11px;font-weight:700;">{f.get('price', '?')}</span>
            </div>
          </td>
        </tr>"""

    # Build flight rows with value badges + v5.0 layover/fare/baggage
    flight_rows = ""
    for i, f in enumerate(flights[:12], 1):
        bg = "#FFFFFF" if i % 2 == 1 else "#F8FAFC"
        price_color = "#059669" if i == 1 else "#1E293B"
        badge = ""
        if i == 1:
            badge = '<span style="background:#059669;color:#fff;font-size:9px;padding:2px 8px;border-radius:12px;margin-left:4px;">CHEAPEST</span>'
        elif f.get("stops", "").lower() == "nonstop" and f == (best_nonstop or {}):
            badge = '<span style="background:#2563EB;color:#fff;font-size:9px;padding:2px 8px;border-radius:12px;margin-left:4px;">BEST NONSTOP</span>'
        val_score = f.get("value_score", 0)
        val_color = "#059669" if val_score >= 85 else "#2563EB" if val_score >= 70 else "#EAB308" if val_score >= 50 else "#F97316" if val_score >= 30 else "#94A3B8"
        stops_color = "#059669" if "nonstop" in f.get("stops", "").lower() else "#64748B"
        booking_url = f.get("booking_link", search_url)

        # v5.0: layover detail line
        detail_parts = []
        if f.get("layover_airport"):
            lay = f"via {f['layover_airport']}"
            if f.get("layover_duration"):
                lay += f" ({f['layover_duration']})"
            detail_parts.append(lay)
        if f.get("fare_class"):
            detail_parts.append(f['fare_class'])
        if f.get("baggage"):
            detail_parts.append(f['baggage'])
        detail_html = ""
        if detail_parts:
            detail_html = f'<div style="color:#94A3B8;font-size:10px;margin-top:3px;">{" · ".join(detail_parts)}</div>'

        flight_rows += f"""
        <tr style="background:{bg};">
          <td style="padding:12px 14px;border-bottom:1px solid #F1F5F9;">
            <div style="font-weight:700;color:{price_color};font-size:16px;">{f.get('price', '—')}{badge}</div>
          </td>
          <td style="padding:12px 14px;border-bottom:1px solid #F1F5F9;">
            <div style="font-weight:600;color:#1E293B;">{f.get('airline', '—')}</div>
            <div style="color:#64748B;font-size:11px;margin-top:2px;">{f.get('duration', '—')}</div>
            {detail_html}
          </td>
          <td style="padding:12px 14px;border-bottom:1px solid #F1F5F9;">
            <div style="font-weight:500;color:#1E293B;font-size:13px;">{f.get('depart_time', '—')} → {f.get('arrive_time', '—')}</div>
            <div style="color:{stops_color};font-size:11px;margin-top:2px;">{f.get('stops', '—')}</div>
          </td>
          <td style="padding:12px 14px;border-bottom:1px solid #F1F5F9;text-align:center;">
            <div style="display:inline-block;background:{val_color};color:#fff;font-size:11px;font-weight:700;padding:3px 8px;border-radius:10px;">{val_score}</div>
          </td>
          <td style="padding:12px 14px;border-bottom:1px solid #F1F5F9;text-align:center;">
            <a href="{booking_url}" style="display:inline-block;background:#0F172A;color:#fff;text-decoration:none;padding:6px 14px;border-radius:6px;font-size:11px;font-weight:600;">Book →</a>
          </td>
        </tr>"""

    # Build suggestions HTML
    suggestions_html = ""
    if suggestions:
        tip_items = ""
        for s in suggestions[:5]:
            tip_items += f"""
            <tr>
              <td style="padding:8px 12px;border-bottom:1px solid #F1F5F9;">
                <div style="font-size:14px;display:inline;margin-right:6px;">{s['icon']}</div>
                <span style="color:#1E293B;font-size:13px;">{s['text']}</span>
              </td>
            </tr>"""
        suggestions_html = f"""
    <div style="margin:0 24px 24px;">
      <h2 style="font-size:16px;color:#0F172A;margin:0 0 12px;font-weight:600;">💡 Smart Suggestions</h2>
      <table width="100%" cellpadding="0" cellspacing="0" style="border-radius:8px;overflow:hidden;border:1px solid #E2E8F0;background:#FFFBEB;">
        {tip_items}
      </table>
    </div>"""

    # Nonstop premium card
    nonstop_card = ""
    if analytics.get("nonstop_premium") is not None and analytics["nonstop_premium"] > 0:
        nonstop_card = f"""
          <td style="width:8px;"></td>
          <td style="background:#F5F3FF;border-radius:8px;padding:14px;text-align:center;width:25%;">
            <div style="font-size:18px;font-weight:700;color:#7C3AED;">+${analytics['nonstop_premium']}</div>
            <div style="color:#6D28D9;font-size:10px;margin-top:2px;">Nonstop Premium</div>
          </td>"""

    cheapest_booking = cheapest.get("booking_link", search_url)

    # Best value pick
    scored = sorted([f for f in flights if f.get("value_score", 0) > 0],
                     key=lambda x: x["value_score"], reverse=True)
    best_val = scored[0] if scored else cheapest
    best_val_card = ""
    if best_val != cheapest:
        best_val_card = f"""
    <div style="margin:0 24px 24px;padding:16px;background:linear-gradient(135deg,#EFF6FF 0%,#DBEAFE 100%);border-radius:10px;border:1px solid #93C5FD;">
      <div style="font-size:11px;text-transform:uppercase;letter-spacing:0.5px;color:#1E40AF;font-weight:600;">⭐ BEST VALUE PICK — Score {best_val['value_score']}/100</div>
      <div style="margin-top:6px;font-size:15px;font-weight:700;color:#1E3A5F;">{best_val.get('price', '?')} · {best_val.get('airline', '?')} · {best_val.get('stops', '?')} · {best_val.get('duration', '?')}</div>
      <div style="margin-top:8px;">
        <a href="{best_val.get('booking_link', search_url)}" style="display:inline-block;background:#2563EB;color:#fff;text-decoration:none;padding:8px 20px;border-radius:6px;font-weight:600;font-size:12px;">Book Best Value →</a>
      </div>
    </div>"""

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#F1F5F9;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;">
  <div style="max-width:680px;margin:0 auto;background:#FFFFFF;">
    <div style="background:linear-gradient(135deg,#0F172A 0%,#1E293B 50%,#0F172A 100%);padding:36px 24px;text-align:center;">
      <div style="font-size:36px;margin-bottom:8px;">✈️</div>
      <h1 style="color:#FFFFFF;font-size:24px;font-weight:700;margin:0;">Flight Intelligence Report</h1>
      <p style="color:#94A3B8;font-size:14px;margin:8px 0 0;">{origin_code} → {dest_code} · {date_display}</p>
      <div style="margin-top:12px;display:inline-block;background:rgba(255,255,255,0.1);padding:4px 14px;border-radius:20px;">
        <span style="color:#94A3B8;font-size:11px;letter-spacing:0.5px;">POWERED BY TARS v6</span>
      </div>
    </div>

    {"" if not price_insight else f'''
    <div style="margin:24px 24px 0;padding:14px 20px;background:linear-gradient(135deg,#FFFBEB 0%,#FEF3C7 100%);border-radius:10px;border:1px solid #FDE68A;">
      <div style="font-size:13px;color:#92400E;">📈 <strong>Google Insight:</strong> {price_insight}</div>
    </div>'''}

    <div style="margin:24px;padding:20px;background:#F8FAFC;border-radius:12px;border:1px solid #E2E8F0;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td style="text-align:center;">
            <div style="font-size:26px;font-weight:700;color:#0F172A;">{origin_code}</div>
            <div style="color:#64748B;font-size:11px;margin-top:2px;">Origin</div>
          </td>
          <td style="text-align:center;padding:0 16px;">
            <div style="color:#94A3B8;font-size:18px;">· · · ✈ · · ·</div>
          </td>
          <td style="text-align:center;">
            <div style="font-size:26px;font-weight:700;color:#0F172A;">{dest_code}</div>
            <div style="color:#64748B;font-size:11px;margin-top:2px;">Destination</div>
          </td>
        </tr>
      </table>
    </div>

    <div style="margin:0 24px 24px;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td style="background:#F0FDF4;border-radius:8px;padding:14px;text-align:center;">
            <div style="font-size:22px;font-weight:700;color:#059669;">{cheapest.get('price', '—')}</div>
            <div style="color:#065F46;font-size:10px;margin-top:2px;">Cheapest</div>
          </td>
          <td style="width:8px;"></td>
          <td style="background:#EFF6FF;border-radius:8px;padding:14px;text-align:center;">
            <div style="font-size:18px;font-weight:700;color:#2563EB;">{analytics['total_flights']}</div>
            <div style="color:#1E40AF;font-size:10px;margin-top:2px;">Flights</div>
          </td>
          <td style="width:8px;"></td>
          <td style="background:#FFF7ED;border-radius:8px;padding:14px;text-align:center;">
            <div style="font-size:18px;font-weight:700;color:#EA580C;">${analytics['price_avg']}</div>
            <div style="color:#9A3412;font-size:10px;margin-top:2px;">Average</div>
          </td>
          <td style="width:8px;"></td>
          <td style="background:#ECFDF5;border-radius:8px;padding:14px;text-align:center;">
            <div style="font-size:18px;font-weight:700;color:#059669;">{best_nonstop.get('price', 'N/A') if best_nonstop else 'N/A'}</div>
            <div style="color:#065F46;font-size:10px;margin-top:2px;">Best Nonstop</div>
          </td>
          {nonstop_card}
        </tr>
      </table>
    </div>

    <div style="margin:0 24px 24px;">
      <h2 style="font-size:14px;color:#64748B;margin:0 0 10px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Price Comparison</h2>
      <div style="background:#F8FAFC;border-radius:8px;padding:16px;border:1px solid #E2E8F0;">
        <table width="100%" cellpadding="0" cellspacing="0">
          {price_bars}
        </table>
      </div>
    </div>

    {best_val_card}

    <div style="margin:0 24px 24px;">
      <h2 style="font-size:14px;color:#64748B;margin:0 0 10px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">All Flights</h2>
      <table width="100%" cellpadding="0" cellspacing="0" style="border-radius:8px;overflow:hidden;border:1px solid #E2E8F0;">
        <tr style="background:#0F172A;">
          <th style="padding:10px 14px;text-align:left;color:#FFFFFF;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Price</th>
          <th style="padding:10px 14px;text-align:left;color:#FFFFFF;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Airline</th>
          <th style="padding:10px 14px;text-align:left;color:#FFFFFF;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Schedule</th>
          <th style="padding:10px 14px;text-align:center;color:#FFFFFF;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Value</th>
          <th style="padding:10px 14px;text-align:center;color:#FFFFFF;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Book</th>
        </tr>
        {flight_rows}
      </table>
      {f'<p style="color:#94A3B8;font-size:11px;margin-top:6px;text-align:center;">Showing top 12 of {len(flights)} results</p>' if len(flights) > 12 else ''}
    </div>

    {suggestions_html}

    {"" if not return_flight or not return_flight.get("depart_time") else f'''
    <div style="margin:0 24px 24px;padding:16px;background:linear-gradient(135deg,#F0F9FF 0%,#E0F2FE 100%);border-radius:10px;border:1px solid #7DD3FC;">
      <div style="font-size:11px;text-transform:uppercase;letter-spacing:0.5px;color:#0369A1;font-weight:600;">🔄 RETURN FLIGHT</div>
      <div style="margin-top:6px;font-size:15px;font-weight:700;color:#0C4A6E;">{return_flight.get("depart_time", "?")} → {return_flight.get("arrive_time", "?")} · {return_flight.get("duration", "?")} · {return_flight.get("stops", "?")}</div>
    </div>'''}

    {"" if not tracker_suggestion else f'''
    <div style="margin:0 24px 24px;padding:14px 20px;background:#F5F3FF;border-radius:10px;border:1px solid #C4B5FD;">
      <div style="font-size:13px;color:#5B21B6;">{tracker_suggestion}</div>
    </div>'''}

    <div style="margin:24px;text-align:center;">
      <a href="{cheapest_booking}" style="display:inline-block;background:#059669;color:#FFFFFF;text-decoration:none;padding:14px 36px;border-radius:8px;font-weight:700;font-size:15px;">🛫 Book Cheapest ({cheapest.get('price', '—')} {cheapest.get('airline', '')}) →</a>
      <div style="margin-top:10px;">
        <a href="{search_url}" style="color:#2563EB;text-decoration:none;font-size:12px;font-weight:500;">View all on Google Flights →</a>
      </div>
    </div>

    <div style="background:#0F172A;padding:24px;text-align:center;">
      <p style="color:#64748B;font-size:11px;margin:0;">TARS Flight Intelligence v6.0 · {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
      <p style="color:#475569;font-size:10px;margin:4px 0 0;">Prices may change. Links verified before sending. Value scores combine price, duration, stops, layover quality, and baggage.</p>
    </div>
  </div>
</body>
</html>"""


def _html_cheapest_dates_email(origin_code, dest_code, date_results, start_date, end_date):
    """Generate a premium HTML email for cheapest dates with analytics, bar chart, and insights."""
    if not date_results:
        return "<p>No results found.</p>"
    cheapest = date_results[0]

    # Compute analytics
    valid_prices = [dr["price_num"] for dr in date_results if dr.get("price_num", 99999) < 99999]
    price_min = min(valid_prices) if valid_prices else 0
    price_avg = round(sum(valid_prices) / len(valid_prices)) if valid_prices else 0
    price_max = max(valid_prices) if valid_prices else 0
    airline_set = set(dr.get("airline", "—") for dr in date_results if dr.get("airline", "—") != "—")

    # Build inline bar chart (CSS bars)
    bar_max = price_max if price_max > 0 else 1
    bar_rows = ""
    sorted_by_date = sorted([dr for dr in date_results if dr.get("price_num", 99999) < 99999],
                            key=lambda x: x["date"])[:20]
    for dr in sorted_by_date:
        pct = min(round((dr["price_num"] / bar_max) * 100), 100)
        bar_color = "#059669" if dr["price_num"] == price_min else "#2563EB" if dr["price_num"] <= price_avg else "#F59E0B" if dr["price_num"] <= price_max * 0.8 else "#DC2626"
        bar_rows += f"""
        <tr>
          <td style="padding:4px 8px;font-size:11px;color:#64748B;white-space:nowrap;width:80px;">{dr['date'][-5:]}</td>
          <td style="padding:4px 8px;width:100%;">
            <div style="background:#F1F5F9;border-radius:4px;height:18px;position:relative;">
              <div style="background:{bar_color};height:18px;border-radius:4px;width:{pct}%;min-width:2px;"></div>
            </div>
          </td>
          <td style="padding:4px 8px;font-size:11px;font-weight:600;color:#0F172A;white-space:nowrap;">{dr['price']}</td>
        </tr>"""

    # Build date table rows
    date_rows = ""
    for i, dr in enumerate(date_results, 1):
        bg = "#FFFFFF" if i % 2 == 1 else "#F8FAFC"
        price_color = "#059669" if i <= 3 else "#1E293B"
        badge = ""
        if i == 1:
            badge = '<span style="background:#059669;color:#fff;font-size:9px;padding:2px 6px;border-radius:10px;margin-left:4px;">🏆 BEST</span>'
        elif i == 2:
            badge = '<span style="background:#EAB308;color:#fff;font-size:9px;padding:2px 6px;border-radius:10px;margin-left:4px;">2nd</span>'
        elif i == 3:
            badge = '<span style="background:#F97316;color:#fff;font-size:9px;padding:2px 6px;border-radius:10px;margin-left:4px;">3rd</span>'
        link = dr.get("booking_link", "#")
        savings_badge = ""
        if i > 1 and dr.get("price_num", 99999) < 99999 and cheapest.get("price_num", 99999) < 99999:
            diff = dr["price_num"] - cheapest["price_num"]
            if diff > 0:
                savings_badge = f'<div style="color:#DC2626;font-size:10px;">+${diff} vs best</div>'
        date_rows += f"""
        <tr style="background:{bg};">
          <td style="padding:12px 14px;border-bottom:1px solid #F1F5F9;">
            <div style="font-weight:600;color:#0F172A;">{dr['date']}</div>
            <div style="color:#94A3B8;font-size:11px;">{dr['day']}</div>
          </td>
          <td style="padding:12px 14px;border-bottom:1px solid #F1F5F9;">
            <div style="font-weight:700;color:{price_color};font-size:15px;">{dr['price']}{badge}</div>
            {savings_badge}
          </td>
          <td style="padding:12px 14px;border-bottom:1px solid #F1F5F9;">
            <div style="color:#1E293B;font-weight:500;">{dr['airline']}</div>
            <div style="color:#94A3B8;font-size:11px;">{dr['stops']} · {dr['duration']}</div>
          </td>
          <td style="padding:12px 14px;border-bottom:1px solid #F1F5F9;text-align:center;">
            <a href="{link}" style="display:inline-block;background:#2563EB;color:#FFFFFF;text-decoration:none;padding:8px 16px;border-radius:8px;font-size:12px;font-weight:600;">Book →</a>
          </td>
        </tr>"""

    # Insight tips
    insights_html = ""
    if len(valid_prices) >= 3:
        price_spread = price_max - price_min
        insights_html += f"""
        <div style="margin:24px;">
          <h2 style="font-size:16px;color:#0F172A;margin:0 0 12px;font-weight:600;">💡 Smart Insights</h2>
          <div style="padding:16px;background:#F0F9FF;border-radius:12px;border:1px solid #BAE6FD;">
            <div style="font-size:13px;color:#0369A1;margin-bottom:8px;">📊 Price range: <strong>${price_min} – ${price_max}</strong> (${price_spread} spread)</div>
            <div style="font-size:13px;color:#0369A1;margin-bottom:8px;">📈 Average price: <strong>${price_avg}</strong> across {len(valid_prices)} dates</div>
            <div style="font-size:13px;color:#0369A1;margin-bottom:8px;">✈️ <strong>{len(airline_set)}</strong> airlines compete on this route</div>
            {"<div style='font-size:13px;color:#059669;'><strong>💰 Flying on " + cheapest.get('day', '') + " saves up to $" + str(price_max - price_min) + " vs the most expensive date!</strong></div>" if price_spread > 50 else ""}
          </div>
        </div>"""

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#F1F5F9;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;">
  <div style="max-width:640px;margin:0 auto;background:#FFFFFF;">
    <div style="background:linear-gradient(135deg,#0F172A 0%,#1E293B 100%);padding:32px 24px;text-align:center;">
      <div style="font-size:32px;margin-bottom:8px;">📊</div>
      <h1 style="color:#FFFFFF;font-size:22px;font-weight:700;margin:0;">Cheapest Dates Report</h1>
      <p style="color:#94A3B8;font-size:14px;margin:8px 0 0;">{origin_code} → {dest_code} · {start_date} – {end_date}</p>
      <p style="color:#475569;font-size:11px;margin:4px 0 0;">TARS Flight Intelligence v6.0</p>
    </div>
    <div style="margin:24px;padding:24px;background:linear-gradient(135deg,#F0FDF4 0%,#DCFCE7 100%);border-radius:12px;border:1px solid #BBF7D0;text-align:center;">
      <div style="font-size:12px;text-transform:uppercase;letter-spacing:1px;color:#065F46;font-weight:600;">🏆 Best Price Found</div>
      <div style="font-size:36px;font-weight:800;color:#059669;margin:8px 0;">{cheapest['price']}</div>
      <div style="color:#065F46;font-size:14px;">{cheapest['date']} ({cheapest['day']}) · {cheapest['airline']} · {cheapest['stops']}</div>
      <a href="{cheapest.get('booking_link', '#')}" style="display:inline-block;background:#059669;color:#FFFFFF;text-decoration:none;padding:12px 28px;border-radius:8px;font-weight:600;font-size:14px;margin-top:12px;">Book This Flight →</a>
    </div>
    <div style="margin:0 24px;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td style="background:#F8FAFC;border-radius:8px;padding:14px;text-align:center;width:25%;">
            <div style="font-size:20px;font-weight:700;color:#059669;">${price_min}</div>
            <div style="color:#64748B;font-size:10px;text-transform:uppercase;">Cheapest</div>
          </td>
          <td style="width:6px;"></td>
          <td style="background:#F8FAFC;border-radius:8px;padding:14px;text-align:center;width:25%;">
            <div style="font-size:20px;font-weight:700;color:#2563EB;">${price_avg}</div>
            <div style="color:#64748B;font-size:10px;text-transform:uppercase;">Average</div>
          </td>
          <td style="width:6px;"></td>
          <td style="background:#F8FAFC;border-radius:8px;padding:14px;text-align:center;width:25%;">
            <div style="font-size:20px;font-weight:700;color:#DC2626;">${price_max}</div>
            <div style="color:#64748B;font-size:10px;text-transform:uppercase;">Highest</div>
          </td>
          <td style="width:6px;"></td>
          <td style="background:#F8FAFC;border-radius:8px;padding:14px;text-align:center;width:25%;">
            <div style="font-size:20px;font-weight:700;color:#7C3AED;">{len(date_results)}</div>
            <div style="color:#64748B;font-size:10px;text-transform:uppercase;">Dates</div>
          </td>
        </tr>
      </table>
    </div>
    <div style="margin:24px;">
      <h2 style="font-size:16px;color:#0F172A;margin:0 0 12px;font-weight:600;">📈 Price by Date</h2>
      <table width="100%" cellpadding="0" cellspacing="0">
        {bar_rows}
      </table>
    </div>
    {insights_html}
    <div style="margin:24px;">
      <h2 style="font-size:16px;color:#0F172A;margin:0 0 12px;font-weight:600;">📋 All Dates (sorted by price)</h2>
      <table width="100%" cellpadding="0" cellspacing="0" style="border-radius:8px;overflow:hidden;border:1px solid #E2E8F0;">
        <tr style="background:#0F172A;">
          <th style="padding:10px 14px;text-align:left;color:#FFFFFF;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Date</th>
          <th style="padding:10px 14px;text-align:left;color:#FFFFFF;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Price</th>
          <th style="padding:10px 14px;text-align:left;color:#FFFFFF;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Details</th>
          <th style="padding:10px 14px;text-align:center;color:#FFFFFF;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">Book</th>
        </tr>
        {date_rows}
      </table>
    </div>
    <div style="background:#F8FAFC;padding:20px 24px;text-align:center;border-top:1px solid #E2E8F0;">
      <p style="color:#94A3B8;font-size:12px;margin:0;">Powered by TARS v6.0 · {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
      <p style="color:#CBD5E1;font-size:11px;margin:4px 0 0;">Prices may change. Book early for best rates. Excel dashboard attached.</p>
    </div>
  </div>
</body>
</html>"""


def _html_price_alert_email(origin_code, dest_code, target_price, current_price,
                            airline, depart_date, booking_link, tracker_id):
    """Generate a beautiful HTML email for price drop alerts."""
    savings = target_price - current_price if current_price < target_price else 0
    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#F1F5F9;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;">
  <div style="max-width:640px;margin:0 auto;background:#FFFFFF;">
    <div style="background:linear-gradient(135deg,#059669 0%,#047857 100%);padding:32px 24px;text-align:center;">
      <div style="font-size:40px;margin-bottom:8px;">🔔</div>
      <h1 style="color:#FFFFFF;font-size:24px;font-weight:700;margin:0;">Price Alert!</h1>
      <p style="color:#A7F3D0;font-size:14px;margin:8px 0 0;">Your tracked route hit your target price</p>
    </div>
    <div style="margin:24px;padding:28px;background:linear-gradient(135deg,#F0FDF4 0%,#DCFCE7 100%);border-radius:16px;border:2px solid #059669;text-align:center;">
      <div style="color:#064E3B;font-size:13px;text-transform:uppercase;letter-spacing:1px;font-weight:600;">Current Price</div>
      <div style="font-size:48px;font-weight:800;color:#059669;margin:8px 0;">${current_price}</div>
      <div style="color:#065F46;font-size:14px;margin-top:4px;">Your target: ${target_price} {'✅' if current_price <= target_price else ''}</div>
      {f'<div style="margin-top:8px;background:#059669;display:inline-block;color:#fff;padding:4px 12px;border-radius:20px;font-size:13px;font-weight:600;">You save ${savings}!</div>' if savings > 0 else ''}
    </div>
    <div style="margin:0 24px 24px;padding:20px;background:#F8FAFC;border-radius:12px;border:1px solid #E2E8F0;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td style="padding:8px 0;">
            <span style="color:#64748B;font-size:12px;">Route</span><br>
            <span style="font-weight:700;color:#0F172A;font-size:16px;">{origin_code} → {dest_code}</span>
          </td>
          <td style="padding:8px 0;text-align:right;">
            <span style="color:#64748B;font-size:12px;">Date</span><br>
            <span style="font-weight:700;color:#0F172A;font-size:16px;">{depart_date}</span>
          </td>
        </tr>
        <tr>
          <td style="padding:8px 0;">
            <span style="color:#64748B;font-size:12px;">Airline</span><br>
            <span style="font-weight:600;color:#0F172A;">{airline}</span>
          </td>
          <td style="padding:8px 0;text-align:right;">
            <span style="color:#64748B;font-size:12px;">Tracker</span><br>
            <span style="font-weight:600;color:#0F172A;">{tracker_id}</span>
          </td>
        </tr>
      </table>
    </div>
    <div style="margin:24px;text-align:center;">
      <a href="{booking_link}" style="display:inline-block;background:#059669;color:#FFFFFF;text-decoration:none;padding:16px 40px;border-radius:10px;font-weight:700;font-size:16px;letter-spacing:0.5px;">🛫 Book Now →</a>
      <p style="color:#94A3B8;font-size:12px;margin-top:12px;">Act fast — prices change frequently!</p>
    </div>
    <div style="background:#F8FAFC;padding:20px 24px;text-align:center;border-top:1px solid #E2E8F0;">
      <p style="color:#94A3B8;font-size:12px;margin:0;">Powered by TARS Flight Tracker · {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
    </div>
  </div>
</body>
</html>"""


# ═══════════════════════════════════════════════════════
#  PHASE 1 — Main Flight Search Function
# ═══════════════════════════════════════════════════════

def _cdp_flight_search(url: str, timeout: int = 50) -> dict:
    """Execute a single CDP-based Google Flights search.

    Returns {"flights": [...], "price_insight": str, "return_flight": dict, "raw_text": str}
    Uses structured DOM extraction first, falls back to text parsing.
    Includes automatic retry on CDP failure.
    """
    for attempt in range(2):  # Retry once on failure
        cdp = None
        try:
            deadline = time.time() + timeout
            cdp = CDP()
            cdp.ensure_connected()
            cdp.send("Page.navigate", {"url": url})
            time.sleep(6)

            if time.time() > deadline:
                raise _FlightSearchTimeout("timeout")

            # Wait for flight results to load
            page_text = ""
            for _ in range(5):
                if time.time() > deadline:
                    raise _FlightSearchTimeout("timeout")
                r = cdp.send("Runtime.evaluate", {
                    "expression": "document.body.innerText.substring(0, 15000)",
                    "returnByValue": True,
                })
                page_text = r.get("result", {}).get("value", "")
                if "$" in page_text and ("stop" in page_text.lower() or "nonstop" in page_text.lower()):
                    break
                time.sleep(2)

            if time.time() > deadline:
                raise _FlightSearchTimeout("timeout")

            # Phase 12: Structured DOM extraction (primary)
            dom_result = _extract_flight_data_dom(cdp)
            flights = dom_result.get("flights", [])
            price_insight = dom_result.get("price_insight", "")
            return_flight = dom_result.get("return_flight", {})

            # If DOM extraction got few results, supplement with text parsing (fallback)
            if len(flights) < 3:
                r2 = cdp.send("Runtime.evaluate", {
                    "expression": """
                        (function() {
                            let results = [];
                            document.querySelectorAll('li, [role="listitem"], [data-resultid]').forEach(el => {
                                let t = el.innerText.trim();
                                if (t.includes('$') && t.length > 20 && t.length < 500) {
                                    results.push(t);
                                }
                            });
                            if (results.length === 0) {
                                let main = document.querySelector('[role="main"]') || document.body;
                                return main.innerText.substring(0, 20000);
                            }
                            return results.join('\\n---\\n');
                        })()
                    """,
                    "returnByValue": True,
                })
                extended_text = r2.get("result", {}).get("value", "")
                combined_text = page_text + "\n" + (extended_text or "")
                fallback_flights = _extract_flight_data(combined_text)
                # Merge: add any fallback flights not already found by DOM
                existing_keys = set()
                for f in flights:
                    existing_keys.add(f"{f.get('price', '')}_{f.get('airline', '')}_{f.get('depart_time', '')}")
                for fb in fallback_flights:
                    key = f"{fb.get('price', '')}_{fb.get('airline', '')}_{fb.get('depart_time', '')}"
                    if key not in existing_keys:
                        flights.append(fb)
                        existing_keys.add(key)

            # Sort by price
            flights.sort(key=lambda f: _price_num(f.get("price", "$99999")))

            if cdp:
                try: cdp.close()
                except Exception: pass

            return {
                "flights": flights,
                "price_insight": price_insight,
                "return_flight": return_flight,
                "raw_text": page_text[:2000],
            }

        except _FlightSearchTimeout:
            if cdp:
                try: cdp.close()
                except Exception: pass
            if attempt == 0:
                logger.warning(f"    ⚠️ CDP attempt {attempt+1} timed out, retrying...")
                time.sleep(2)
                continue
            return {"flights": [], "price_insight": "", "return_flight": {}, "raw_text": ""}

        except Exception as e:
            if cdp:
                try: cdp.close()
                except Exception: pass
            if attempt == 0:
                logger.warning(f"    ⚠️ CDP attempt {attempt+1} failed ({e}), retrying...")
                time.sleep(2)
                continue
            return {"flights": [], "price_insight": "", "return_flight": {}, "raw_text": ""}

    return {"flights": [], "price_insight": "", "return_flight": {}, "raw_text": ""}


def search_flights(
    origin: str, destination: str, depart_date: str,
    return_date: str = "", passengers: int = 1,
    trip_type: str = "round_trip", cabin: str = "economy",
    stops: str = "any", sort_by: str = "price", max_price: int = 0,
) -> dict:
    """Search for flights using Google Flights (v5.0 — structured DOM + cache + retry)."""
    try:
        origin_code = _resolve_airport(origin)
        dest_code = _resolve_airport(destination)
        depart = _parse_date(depart_date)
        if trip_type == "one_way":
            return_date = ""
        ret_parsed = _parse_date(return_date) if return_date else ""

        query_desc = f"{origin_code} → {dest_code} on {depart}"
        if ret_parsed:
            query_desc += f", returning {ret_parsed}"
        if stops != "any":
            query_desc += f" ({stops})"
        if cabin != "economy":
            query_desc += f" [{cabin}]"

        # Phase 15: Check cache first
        c_key = _cache_key(origin_code, dest_code, depart, ret_parsed, cabin, stops)
        cached = _get_cached(c_key)
        if cached:
            logger.info(f"    ⚡ Cache hit for {origin_code}→{dest_code}")
            return cached

        url = _build_google_flights_url(
            origin=origin, destination=destination,
            depart_date=depart_date, return_date=return_date,
            passengers=passengers, trip_type=trip_type,
            cabin=cabin, stops=stops, sort_by=sort_by,
        )

        # Phase 12+15: Structured DOM search with CDP retry
        cdp_result = _cdp_flight_search(url, timeout=50)
        flights = cdp_result.get("flights", [])
        price_insight = cdp_result.get("price_insight", "")
        return_flight = cdp_result.get("return_flight", {})

        if max_price > 0:
            flights = [f for f in flights if _price_num(f.get("price", "$99999")) <= max_price]

        # Assign per-flight airline booking links (verified)
        for f in flights:
            airline = f.get("airline", "—")
            verified_url = _get_verified_booking_link(
                airline, origin, destination, depart_date, return_date, search_url=url)
            f["booking_link"] = verified_url

        # v5.0 — Intelligence layer (value scores + analytics)
        intel = _analyze_flights(flights, origin_code, dest_code, depart, ret_parsed)

        # Auto-suggest tracker target
        tracker_suggestion = ""
        if intel["analytics"].get("price_avg") and intel["analytics"].get("price_min"):
            avg = intel["analytics"]["price_avg"]
            mn = intel["analytics"]["price_min"]
            suggested_target = round(mn * 0.85 / 10) * 10  # 15% below current cheapest, rounded to $10
            if suggested_target > 50:
                tracker_suggestion = f"💡 Set a price tracker at ${suggested_target} (15% below current cheapest) to catch a deal."

        report = _format_flights_rich(flights, query_desc, origin_code, dest_code, depart, ret_parsed,
                                       price_insight=price_insight, return_flight=return_flight,
                                       tracker_suggestion=tracker_suggestion)

        result = {
            "success": True, "content": report, "flights": flights,
            "url": url, "source": "Google Flights",
            "analytics": intel["analytics"], "suggestions": intel["suggestions"],
            "price_insight": price_insight,
            "return_flight": return_flight,
            "tracker_suggestion": tracker_suggestion,
        }

        # Cache the result
        _set_cache(c_key, result)

        return result
    except Exception as e:
        return {"success": False, "content": f"❌ Flight search failed: {str(e)}", "flights": [], "error": str(e)}


# ═══════════════════════════════════════════════════════
#  PHASE 5 — Enhanced Excel with Hyperlinks
# ═══════════════════════════════════════════════════════

def _generate_flight_excel(title, flights, origin_code, dest_code, search_url, summary_data=None, analytics=None, suggestions=None):
    """Generate a premium Excel report with charts, conditional formatting, hyperlinks, and dashboard."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.chart.label import DataLabelList
        from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
    except ImportError:
        return {"success": False, "content": "openpyxl not installed"}

    REPORT_DIR = os.path.expanduser("~/Documents/TARS_Reports")
    os.makedirs(REPORT_DIR, exist_ok=True)
    wb = Workbook()

    # ── Color Palette ──
    DARK, GREEN, BLUE, GRAY = "0F172A", "059669", "2563EB", "64748B"
    ALT_ROW, WHITE, RED = "F1F5F9", "FFFFFF", "DC2626"
    GOLD, ORANGE, PURPLE = "F59E0B", "EA580C", "7C3AED"
    header_font = Font(name="Helvetica Neue", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
    data_font = Font(name="Helvetica Neue", size=10)
    link_font = Font(name="Helvetica Neue", size=10, color=BLUE, underline="single")
    alt_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    gold_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0"),
    )

    # ═══════════════════════════════════════════════════
    #  SHEET 1: Dashboard Overview
    # ═══════════════════════════════════════════════════
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_properties.tabColor = "059669"

    # Title
    ws_dash.merge_cells("A1:H1")
    ws_dash.cell(row=1, column=1, value=f"✈️ TARS Flight Intelligence Dashboard").font = Font(
        name="Helvetica Neue", size=22, bold=True, color=DARK)
    ws_dash.row_dimensions[1].height = 48
    ws_dash.merge_cells("A2:H2")
    ws_dash.cell(row=2, column=1, value=f"{title} · Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}").font = Font(
        name="Helvetica Neue", size=10, italic=True, color=GRAY)
    ws_dash.row_dimensions[2].height = 22

    # KPI Cards Row
    kpi_row = 4
    kpi_items = []
    if analytics:
        kpi_items = [
            ("💰 CHEAPEST", f"${analytics.get('price_min', 'N/A')}", GREEN),
            ("📊 AVERAGE", f"${analytics.get('price_avg', 'N/A')}", BLUE),
            ("📈 HIGHEST", f"${analytics.get('price_max', 'N/A')}", RED),
            ("✈️ TOTAL", str(analytics.get('total_flights', len(flights))), DARK),
            ("🛫 NONSTOP", str(analytics.get('nonstop_count', 0)), PURPLE),
            ("🏢 AIRLINES", str(analytics.get('airline_count', 0)), ORANGE),
        ]
    else:
        cheapest_price = flights[0].get("price", "?") if flights else "?"
        kpi_items = [
            ("💰 CHEAPEST", str(cheapest_price), GREEN),
            ("✈️ TOTAL", str(len(flights)), DARK),
        ]

    for ki, (kpi_label, kpi_val, kpi_color) in enumerate(kpi_items):
        col = ki * 2 + 1
        # Card background
        for r_off in range(3):
            for c_off in range(2):
                cell = ws_dash.cell(row=kpi_row + r_off, column=col + c_off)
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                cell.border = thin_border
        ws_dash.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=col + 1)
        ws_dash.cell(row=kpi_row, column=col, value=kpi_label).font = Font(
            name="Helvetica Neue", size=9, bold=True, color=GRAY)
        ws_dash.cell(row=kpi_row, column=col).alignment = Alignment(horizontal="center")
        ws_dash.merge_cells(start_row=kpi_row + 1, start_column=col, end_row=kpi_row + 1, end_column=col + 1)
        ws_dash.cell(row=kpi_row + 1, column=col, value=kpi_val).font = Font(
            name="Helvetica Neue", size=20, bold=True, color=kpi_color)
        ws_dash.cell(row=kpi_row + 1, column=col).alignment = Alignment(horizontal="center")

    # ── Price Bar Chart ──
    chart_data_row = kpi_row + 5
    ws_dash.cell(row=chart_data_row - 1, column=1, value="Price Chart Data:").font = Font(
        name="Helvetica Neue", size=8, color=GRAY)
    ws_dash.cell(row=chart_data_row, column=1, value="Airline").font = Font(
        name="Helvetica Neue", size=9, bold=True, color=GRAY)
    ws_dash.cell(row=chart_data_row, column=2, value="Price ($)").font = Font(
        name="Helvetica Neue", size=9, bold=True, color=GRAY)

    chart_flights = flights[:10]
    for ci, f in enumerate(chart_flights):
        ws_dash.cell(row=chart_data_row + 1 + ci, column=1, value=f.get("airline", "?")[:15]).font = data_font
        price_val = _price_num(f.get("price", "$0"))
        ws_dash.cell(row=chart_data_row + 1 + ci, column=2, value=price_val).font = data_font

    if chart_flights:
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Price Comparison"
        bar_chart.y_axis.title = "Price ($)"
        bar_chart.x_axis.title = "Airline"
        bar_chart.width = 24
        bar_chart.height = 14

        data_ref = Reference(ws_dash, min_col=2, min_row=chart_data_row,
                             max_row=chart_data_row + len(chart_flights))
        cats_ref = Reference(ws_dash, min_col=1, min_row=chart_data_row + 1,
                             max_row=chart_data_row + len(chart_flights))
        bar_chart.add_data(data_ref, titles_from_data=True)
        bar_chart.set_categories(cats_ref)
        bar_chart.shape = 4

        # Color the bars
        from openpyxl.chart.series import DataPoint
        from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
        series = bar_chart.series[0]
        series.graphicalProperties.solidFill = "2563EB"
        # Highlight cheapest bar in green
        pt = DataPoint(idx=0)
        pt.graphicalProperties.solidFill = "059669"
        series.data_points.append(pt)

        ws_dash.add_chart(bar_chart, f"D{chart_data_row - 1}")

    # ── Airline Pie Chart ──
    if analytics and analytics.get("airline_stats"):
        airline_stats = analytics["airline_stats"]
        pie_data_row = chart_data_row + len(chart_flights) + 4
        ws_dash.cell(row=pie_data_row - 1, column=1, value="Airline Distribution:").font = Font(
            name="Helvetica Neue", size=8, color=GRAY)
        ws_dash.cell(row=pie_data_row, column=1, value="Airline").font = Font(
            name="Helvetica Neue", size=9, bold=True, color=GRAY)
        ws_dash.cell(row=pie_data_row, column=2, value="Flights").font = Font(
            name="Helvetica Neue", size=9, bold=True, color=GRAY)

        for ai, (aname, adata) in enumerate(list(airline_stats.items())[:8]):
            ws_dash.cell(row=pie_data_row + 1 + ai, column=1, value=aname).font = data_font
            ws_dash.cell(row=pie_data_row + 1 + ai, column=2, value=adata.get("count", 0)).font = data_font

        if airline_stats:
            pie_chart = PieChart()
            pie_chart.title = "Flights by Airline"
            pie_chart.style = 10
            pie_chart.width = 14
            pie_chart.height = 12

            pie_data = Reference(ws_dash, min_col=2, min_row=pie_data_row,
                                 max_row=pie_data_row + min(len(airline_stats), 8))
            pie_cats = Reference(ws_dash, min_col=1, min_row=pie_data_row + 1,
                                 max_row=pie_data_row + min(len(airline_stats), 8))
            pie_chart.add_data(pie_data, titles_from_data=True)
            pie_chart.set_categories(pie_cats)

            # Labels
            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showPercent = True
            pie_chart.dataLabels.showVal = False

            # Color slices
            pie_colors = ["059669", "2563EB", "F59E0B", "DC2626", "7C3AED", "EA580C", "0891B2", "4F46E5"]
            for pi in range(min(len(airline_stats), 8)):
                pt = DataPoint(idx=pi)
                pt.graphicalProperties.solidFill = pie_colors[pi % len(pie_colors)]
                pie_chart.series[0].data_points.append(pt)

            ws_dash.add_chart(pie_chart, f"D{pie_data_row - 1}")

    # ── Summary Info ──
    if summary_data:
        srow = 4
        scol = 14  # Far right
        ws_dash.cell(row=srow, column=scol, value="📋 SEARCH SUMMARY").font = Font(
            name="Helvetica Neue", size=12, bold=True, color=DARK)
        for si, (sk, sv) in enumerate(summary_data.items()):
            ws_dash.cell(row=srow + 1 + si, column=scol, value=sk).font = Font(
                name="Helvetica Neue", size=10, bold=True, color=GRAY)
            ws_dash.cell(row=srow + 1 + si, column=scol + 1, value=sv).font = data_font

    # ── Suggestions ──
    if suggestions:
        sug_row = (srow + len(summary_data) + 3) if summary_data else kpi_row + 8
        ws_dash.cell(row=sug_row, column=14, value="💡 SMART TIPS").font = Font(
            name="Helvetica Neue", size=12, bold=True, color=DARK)
        for si, s in enumerate(suggestions[:6]):
            ws_dash.cell(row=sug_row + 1 + si, column=14,
                         value=f"{s['icon']} {s['text']}").font = Font(
                name="Helvetica Neue", size=10, color=DARK)

    # Column widths for dashboard
    for col in range(1, 16):
        ws_dash.column_dimensions[get_column_letter(col)].width = 12

    # ═══════════════════════════════════════════════════
    #  SHEET 2: All Flights (the data table)
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet(title="Flights")
    ws.sheet_properties.tabColor = "2563EB"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(row=1, column=1, value=f"✈️ {title}").font = Font(name="Helvetica Neue", size=18, bold=True, color=DARK)
    ws.row_dimensions[1].height = 40
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.cell(row=2, column=1, value=f"Generated by TARS · {datetime.now().strftime('%B %d, %Y at %I:%M %p')}").font = Font(
        name="Helvetica Neue", size=9, italic=True, color=GRAY)

    headers = ["#", "Price", "Airline", "Departure", "Arrival", "Duration", "Stops", "Layover", "Fare", "Baggage", "Value", "Book"]
    start_row = 4
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[start_row].height = 30

    for row_idx, f in enumerate(flights[:20], start_row + 1):
        i = row_idx - start_row
        val_score = f.get("value_score", 0)
        # Build layover display
        layover = ""
        if f.get("layover_airport"):
            layover = f"via {f['layover_airport']}"
            if f.get("layover_duration"):
                layover += f" ({f['layover_duration']})"

        price_val = _price_num(f.get("price", "$0"))
        values = [i, price_val, f.get("airline", "—"), f.get("depart_time", "—"),
                  f.get("arrive_time", "—"), f.get("duration", "—"), f.get("stops", "—"),
                  layover, f.get("fare_class", "—"), f.get("baggage", "—"), val_score, ""]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if i % 2 == 0:
                cell.fill = alt_fill
            if col_idx == 2:  # Price column
                cell.number_format = '$#,##0'
                if i == 1:
                    cell.font = Font(name="Helvetica Neue", size=11, bold=True, color=GREEN)
                    cell.fill = green_fill
                else:
                    cell.font = Font(name="Helvetica Neue", size=10, bold=True)
            if col_idx == 11:  # Value column
                if val_score >= 80:
                    cell.font = Font(name="Helvetica Neue", size=10, bold=True, color=GREEN)
                elif val_score >= 60:
                    cell.font = Font(name="Helvetica Neue", size=10, bold=True, color=BLUE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        link_cell = ws.cell(row=row_idx, column=12, value="View →")
        link_cell.hyperlink = f.get("booking_link", search_url)
        link_cell.font = link_font
        link_cell.alignment = Alignment(horizontal="center", vertical="center")
        link_cell.border = thin_border

    # ── Conditional Formatting: Price column data bars ──
    data_end_row = start_row + min(len(flights), 20)
    price_range_str = f"B{start_row+1}:B{data_end_row}"
    ws.conditional_formatting.add(price_range_str, DataBarRule(
        start_type="min", end_type="max",
        color="2563EB", showValue=True, minLength=None, maxLength=None,
    ))

    # ── Conditional Formatting: Value score color scale ──
    value_range_str = f"K{start_row+1}:K{data_end_row}"
    ws.conditional_formatting.add(value_range_str, ColorScaleRule(
        start_type="min", start_color="FEE2E2",  # Red for low
        mid_type="percentile", mid_value=50, mid_color="FEF3C7",  # Yellow for mid
        end_type="max", end_color="DCFCE7",  # Green for high
    ))

    # ── Conditional Formatting: Highlight cheapest row ──
    ws.conditional_formatting.add(f"A{start_row+1}:L{start_row+1}", CellIsRule(
        operator="notEqual", formula=['""'],
        fill=green_fill,
    ))

    if summary_data:
        srow = start_row + len(flights[:20]) + 2
        ws.cell(row=srow, column=1, value="Summary").font = Font(name="Helvetica Neue", size=13, bold=True, color=DARK)
        for i, (k, v) in enumerate(summary_data.items()):
            ws.cell(row=srow + 1 + i, column=1, value=k).font = Font(name="Helvetica Neue", size=10, bold=True)
            ws.cell(row=srow + 1 + i, column=2, value=v).font = data_font

    widths = [5, 10, 18, 12, 12, 14, 12, 18, 16, 18, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════════════════
    #  SHEET 3: Insights (analytics + suggestions)
    # ═══════════════════════════════════════════════════
    if analytics or suggestions:
        ws2 = wb.create_sheet(title="Insights")
        ws2.sheet_properties.tabColor = "F59E0B"
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws2.cell(row=1, column=1, value="💡 Flight Intelligence Insights").font = Font(name="Helvetica Neue", size=16, bold=True, color=DARK)
        ws2.row_dimensions[1].height = 36

        if analytics:
            ws2.cell(row=3, column=1, value="📊 PRICE ANALYTICS").font = Font(name="Helvetica Neue", size=12, bold=True, color=DARK)
            stat_keys = [("Total Flights", "total_flights"), ("Cheapest", "price_min"), ("Average", "price_avg"),
                         ("Highest", "price_max"), ("Median", "price_median"), ("Std Dev", "price_std_dev"),
                         ("Range", "price_range"), ("Nonstop Premium", "nonstop_premium")]
            for i, (label, key) in enumerate(stat_keys):
                ws2.cell(row=4+i, column=1, value=label).font = Font(name="Helvetica Neue", size=10, bold=True)
                val = analytics.get(key, "N/A")
                if isinstance(val, (int, float)) and key not in ("total_flights",):
                    val = f"${val}"
                ws2.cell(row=4+i, column=2, value=str(val)).font = data_font

            # Airline breakdown with mini price bars
            airlines = analytics.get("airline_stats", analytics.get("airlines", {}))
            if airlines:
                arow = 4 + len(stat_keys) + 2
                ws2.cell(row=arow, column=1, value="✈️ AIRLINE COMPARISON").font = Font(name="Helvetica Neue", size=12, bold=True, color=DARK)
                arow += 1
                for col_idx, h in enumerate(["Airline", "Cheapest", "Average", "Flights", "Nonstop", "Status"], 1):
                    cell = ws2.cell(row=arow, column=col_idx, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                for ai, (name, data) in enumerate(airlines.items()):
                    r = arow + 1 + ai
                    ws2.cell(row=r, column=1, value=name).font = Font(name="Helvetica Neue", size=10, bold=True)
                    min_val = data.get('min', 0)
                    avg_val = data.get('avg', 0)
                    ws2.cell(row=r, column=2, value=min_val).font = data_font
                    ws2.cell(row=r, column=2).number_format = '$#,##0'
                    ws2.cell(row=r, column=3, value=round(avg_val)).font = data_font
                    ws2.cell(row=r, column=3).number_format = '$#,##0'
                    ws2.cell(row=r, column=4, value=data.get('count', 0)).font = data_font
                    has_ns = data.get('has_nonstop', data.get('nonstop', 0))
                    ws2.cell(row=r, column=5, value="✓" if has_ns else "—").font = data_font
                    # Status: cheapest airline gets green, expensive gets red
                    if ai == 0:
                        ws2.cell(row=r, column=6, value="💰 Best Price").font = Font(
                            name="Helvetica Neue", size=10, bold=True, color=GREEN)
                    elif min_val and analytics.get("price_avg") and min_val > analytics["price_avg"]:
                        ws2.cell(row=r, column=6, value="Above Average").font = Font(
                            name="Helvetica Neue", size=10, color=RED)
                    else:
                        ws2.cell(row=r, column=6, value="Competitive").font = Font(
                            name="Helvetica Neue", size=10, color=BLUE)

                # Add data bars to Cheapest column
                cheapest_range = f"B{arow+1}:B{arow+len(airlines)}"
                ws2.conditional_formatting.add(cheapest_range, DataBarRule(
                    start_type="min", end_type="max",
                    color="059669", showValue=True, minLength=None, maxLength=None,
                ))

                # ── Airline Price Comparison Chart ──
                airline_chart = BarChart()
                airline_chart.type = "bar"
                airline_chart.style = 10
                airline_chart.title = "Cheapest Price by Airline"
                airline_chart.x_axis.title = "Price ($)"
                airline_chart.width = 20
                airline_chart.height = max(8, len(airlines))

                ac_data = Reference(ws2, min_col=2, min_row=arow,
                                    max_row=arow + len(airlines))
                ac_cats = Reference(ws2, min_col=1, min_row=arow + 1,
                                    max_row=arow + len(airlines))
                airline_chart.add_data(ac_data, titles_from_data=True)
                airline_chart.set_categories(ac_cats)
                airline_chart.series[0].graphicalProperties.solidFill = "2563EB"

                ws2.add_chart(airline_chart, f"A{arow + len(airlines) + 2}")

        if suggestions:
            srow2 = (4 + len(stat_keys) + 2 + len(analytics.get("airline_stats", analytics.get("airlines", {}))) + 4) if analytics else 3
            srow2 = max(srow2, 3)
            # Find the right row after any chart
            if analytics and analytics.get("airline_stats", analytics.get("airlines", {})):
                srow2 = 4 + len(stat_keys) + 2 + len(analytics.get("airline_stats", analytics.get("airlines", {}))) + 20  # After chart
            ws2.cell(row=srow2, column=1, value="💡 SMART SUGGESTIONS").font = Font(name="Helvetica Neue", size=12, bold=True, color=DARK)
            for si, s in enumerate(suggestions[:8]):
                ws2.cell(row=srow2+1+si, column=1, value=f"{s['icon']} {s['type'].upper()}").font = Font(name="Helvetica Neue", size=10, bold=True)
                ws2.merge_cells(start_row=srow2+1+si, start_column=2, end_row=srow2+1+si, end_column=5)
                ws2.cell(row=srow2+1+si, column=2, value=s['text']).font = data_font

        for col in range(1, 7):
            ws2.column_dimensions[get_column_letter(col)].width = 20

    safe_title = "".join(c if c.isalnum() or c in " _-" else "_" for c in title)
    filename = f"{safe_title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(REPORT_DIR, filename)
    wb.save(filepath)
    return {"success": True, "path": filepath}


def _generate_dates_excel(origin_code, dest_code, date_results, start, end):
    """Generate a premium Excel dashboard for cheapest dates with charts, data bars, and conditional formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.chart.label import DataLabelList
        from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
    except ImportError:
        return {"success": False, "content": "openpyxl not installed"}

    REPORT_DIR = os.path.expanduser("~/Documents/TARS_Reports")
    os.makedirs(REPORT_DIR, exist_ok=True)
    wb = Workbook()

    # ── Color Palette ──
    DARK, GREEN, BLUE, GRAY = "0F172A", "059669", "2563EB", "64748B"
    ALT_ROW, WHITE, RED = "F1F5F9", "FFFFFF", "DC2626"
    GOLD, ORANGE, PURPLE = "F59E0B", "EA580C", "7C3AED"
    header_font = Font(name="Helvetica Neue", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
    data_font = Font(name="Helvetica Neue", size=10)
    link_font = Font(name="Helvetica Neue", size=10, color=BLUE, underline="single")
    gold_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    alt_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0"),
    )

    # Compute analytics from date_results
    valid_prices = [dr["price_num"] for dr in date_results if dr.get("price_num", 99999) < 99999]
    cheapest = date_results[0]
    airline_counts = {}
    for dr in date_results:
        a = dr.get("airline", "—")
        if a != "—":
            if a not in airline_counts:
                airline_counts[a] = {"count": 0, "min_price": 99999, "prices": []}
            airline_counts[a]["count"] += 1
            p = dr.get("price_num", 99999)
            if p < 99999:
                airline_counts[a]["prices"].append(p)
                airline_counts[a]["min_price"] = min(airline_counts[a]["min_price"], p)
    for a in airline_counts:
        prices = airline_counts[a]["prices"]
        airline_counts[a]["avg_price"] = round(sum(prices) / len(prices)) if prices else 0

    # ═══════════════════════════════════════════════════
    #  SHEET 1: Dashboard Overview
    # ═══════════════════════════════════════════════════
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_properties.tabColor = "059669"

    # Title
    ws_dash.merge_cells("A1:J1")
    ws_dash.cell(row=1, column=1, value=f"📊 TARS Cheapest Dates Dashboard").font = Font(
        name="Helvetica Neue", size=22, bold=True, color=DARK)
    ws_dash.row_dimensions[1].height = 48
    ws_dash.merge_cells("A2:J2")
    ws_dash.cell(row=2, column=1, value=f"{origin_code} → {dest_code} · {start.strftime('%b %d')} – {end.strftime('%b %d, %Y')} · Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}").font = Font(
        name="Helvetica Neue", size=10, italic=True, color=GRAY)
    ws_dash.row_dimensions[2].height = 22

    # KPI Cards Row
    kpi_row = 4
    price_min = min(valid_prices) if valid_prices else 0
    price_avg = round(sum(valid_prices) / len(valid_prices)) if valid_prices else 0
    price_max = max(valid_prices) if valid_prices else 0
    kpi_items = [
        ("💰 CHEAPEST", f"${price_min}", GREEN),
        ("📊 AVERAGE", f"${price_avg}", BLUE),
        ("📈 HIGHEST", f"${price_max}", RED),
        ("📅 DATES", str(len(date_results)), DARK),
        ("✈️ AIRLINES", str(len(airline_counts)), PURPLE),
        ("✅ WITH FLIGHTS", str(len(valid_prices)), ORANGE),
    ]

    for ki, (kpi_label, kpi_val, kpi_color) in enumerate(kpi_items):
        col = ki * 2 + 1
        for r_off in range(3):
            for c_off in range(2):
                cell = ws_dash.cell(row=kpi_row + r_off, column=col + c_off)
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                cell.border = thin_border
        ws_dash.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=col + 1)
        ws_dash.cell(row=kpi_row, column=col, value=kpi_label).font = Font(
            name="Helvetica Neue", size=9, bold=True, color=GRAY)
        ws_dash.cell(row=kpi_row, column=col).alignment = Alignment(horizontal="center")
        ws_dash.merge_cells(start_row=kpi_row + 1, start_column=col, end_row=kpi_row + 1, end_column=col + 1)
        ws_dash.cell(row=kpi_row + 1, column=col, value=kpi_val).font = Font(
            name="Helvetica Neue", size=20, bold=True, color=kpi_color)
        ws_dash.cell(row=kpi_row + 1, column=col).alignment = Alignment(horizontal="center")

    # ── Price Timeline Chart ──
    chart_data_row = kpi_row + 5
    ws_dash.cell(row=chart_data_row - 1, column=1, value="Price by Date:").font = Font(
        name="Helvetica Neue", size=8, color=GRAY)
    ws_dash.cell(row=chart_data_row, column=1, value="Date").font = Font(
        name="Helvetica Neue", size=9, bold=True, color=GRAY)
    ws_dash.cell(row=chart_data_row, column=2, value="Price ($)").font = Font(
        name="Helvetica Neue", size=9, bold=True, color=GRAY)

    # Sort by date for the timeline
    sorted_by_date = sorted([dr for dr in date_results if dr.get("price_num", 99999) < 99999],
                            key=lambda x: x["date"])
    for ci, dr in enumerate(sorted_by_date[:30]):
        ws_dash.cell(row=chart_data_row + 1 + ci, column=1, value=dr["date"]).font = data_font
        ws_dash.cell(row=chart_data_row + 1 + ci, column=2, value=dr["price_num"]).font = data_font

    if sorted_by_date:
        line_chart = LineChart()
        line_chart.style = 10
        line_chart.title = "Price by Travel Date"
        line_chart.y_axis.title = "Price ($)"
        line_chart.x_axis.title = "Date"
        line_chart.width = 24
        line_chart.height = 14

        data_ref = Reference(ws_dash, min_col=2, min_row=chart_data_row,
                             max_row=chart_data_row + len(sorted_by_date[:30]))
        cats_ref = Reference(ws_dash, min_col=1, min_row=chart_data_row + 1,
                             max_row=chart_data_row + len(sorted_by_date[:30]))
        line_chart.add_data(data_ref, titles_from_data=True)
        line_chart.set_categories(cats_ref)
        line_chart.series[0].graphicalProperties.line.solidFill = "2563EB"
        line_chart.series[0].graphicalProperties.line.width = 25000

        ws_dash.add_chart(line_chart, f"D{chart_data_row - 1}")

    # ── Airline Price Comparison Bar Chart ──
    if airline_counts:
        airline_chart_row = chart_data_row + max(len(sorted_by_date[:30]), 16) + 3
        ws_dash.cell(row=airline_chart_row - 1, column=1, value="Airline Comparison:").font = Font(
            name="Helvetica Neue", size=8, color=GRAY)
        ws_dash.cell(row=airline_chart_row, column=1, value="Airline").font = Font(
            name="Helvetica Neue", size=9, bold=True, color=GRAY)
        ws_dash.cell(row=airline_chart_row, column=2, value="Cheapest ($)").font = Font(
            name="Helvetica Neue", size=9, bold=True, color=GRAY)

        sorted_airlines = sorted(airline_counts.items(), key=lambda x: x[1]["min_price"])
        for ai, (aname, adata) in enumerate(sorted_airlines[:8]):
            ws_dash.cell(row=airline_chart_row + 1 + ai, column=1, value=aname[:15]).font = data_font
            ws_dash.cell(row=airline_chart_row + 1 + ai, column=2, value=adata["min_price"]).font = data_font

        bar_chart = BarChart()
        bar_chart.type = "bar"
        bar_chart.style = 10
        bar_chart.title = "Cheapest Price by Airline"
        bar_chart.x_axis.title = "Price ($)"
        bar_chart.width = 24
        bar_chart.height = 12

        data_ref = Reference(ws_dash, min_col=2, min_row=airline_chart_row,
                             max_row=airline_chart_row + len(sorted_airlines[:8]))
        cats_ref = Reference(ws_dash, min_col=1, min_row=airline_chart_row + 1,
                             max_row=airline_chart_row + len(sorted_airlines[:8]))
        bar_chart.add_data(data_ref, titles_from_data=True)
        bar_chart.set_categories(cats_ref)
        bar_chart.series[0].graphicalProperties.solidFill = "2563EB"
        pt = DataPoint(idx=0)
        pt.graphicalProperties.solidFill = "059669"
        bar_chart.series[0].data_points.append(pt)

        ws_dash.add_chart(bar_chart, f"D{airline_chart_row - 1}")

    # ── Best Date Box ──
    best_row = kpi_row + 3
    ws_dash.cell(row=best_row, column=1, value="🏆 Best Date:").font = Font(
        name="Helvetica Neue", size=10, bold=True, color=GRAY)
    ws_dash.cell(row=best_row, column=2, value=f"{cheapest['date']} ({cheapest['day']}) — {cheapest['price']} on {cheapest['airline']}").font = Font(
        name="Helvetica Neue", size=10, bold=True, color=GREEN)

    for col in range(1, 14):
        ws_dash.column_dimensions[get_column_letter(col)].width = 12

    # ═══════════════════════════════════════════════════
    #  SHEET 2: All Dates (data table)
    # ═══════════════════════════════════════════════════
    ws = wb.create_sheet(title="All Dates")
    ws.sheet_properties.tabColor = "2563EB"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=1, column=1, value=f"📊 Cheapest Dates: {origin_code} → {dest_code}").font = Font(
        name="Helvetica Neue", size=18, bold=True, color=DARK)
    ws.row_dimensions[1].height = 40
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.cell(row=2, column=1, value=f"Range: {start.strftime('%b %d')} – {end.strftime('%b %d, %Y')} · Generated by TARS v6.0 · {datetime.now().strftime('%B %d, %Y')}").font = Font(
        name="Helvetica Neue", size=9, italic=True, color=GRAY)

    headers = ["#", "Date", "Day", "Price ($)", "Airline", "Departure", "Duration", "Stops", "Options", "Book"]
    data_start = 4
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[data_start].height = 30

    for row_idx, dr in enumerate(date_results, data_start + 1):
        i = row_idx - data_start
        price_val = dr.get("price_num", 99999) if dr.get("price_num", 99999) < 99999 else None
        values = [i, dr["date"], dr["day"], price_val, dr["airline"],
                  dr["depart_time"], dr["duration"], dr["stops"], dr["options"], ""]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if i == 1:
                cell.fill = green_fill
            elif i <= 3:
                cell.fill = gold_fill
            elif i % 2 == 0:
                cell.fill = alt_fill
            if col_idx == 4:  # Price column
                cell.number_format = '$#,##0'
                if i == 1:
                    cell.font = Font(name="Helvetica Neue", size=11, bold=True, color=GREEN)
                elif i <= 3:
                    cell.font = Font(name="Helvetica Neue", size=10, bold=True, color=GOLD)
                else:
                    cell.font = Font(name="Helvetica Neue", size=10, bold=True)
        link_cell = ws.cell(row=row_idx, column=10, value="Book →")
        link_cell.hyperlink = dr.get("booking_link", "")
        link_cell.font = link_font
        link_cell.alignment = Alignment(horizontal="center", vertical="center")
        link_cell.border = thin_border

    # ── Conditional Formatting: Price data bars ──
    data_end_row = data_start + len(date_results)
    price_range_str = f"D{data_start+1}:D{data_end_row}"
    ws.conditional_formatting.add(price_range_str, DataBarRule(
        start_type="min", end_type="max",
        color="2563EB", showValue=True, minLength=None, maxLength=None,
    ))

    # ── Conditional Formatting: Price color scale ──
    ws.conditional_formatting.add(price_range_str, ColorScaleRule(
        start_type="min", start_color="DCFCE7",
        mid_type="percentile", mid_value=50, mid_color="FEF3C7",
        end_type="max", end_color="FEE2E2",
    ))

    # ── Conditional Formatting: Highlight cheapest row ──
    ws.conditional_formatting.add(f"A{data_start+1}:J{data_start+1}", CellIsRule(
        operator="notEqual", formula=['""'],
        fill=green_fill,
    ))

    # Summary below data
    srow = data_start + len(date_results) + 2
    ws.cell(row=srow, column=1, value="📋 Summary").font = Font(name="Helvetica Neue", size=13, bold=True, color=DARK)
    summary = {
        "Route": f"{origin_code} → {dest_code}",
        "Dates Scanned": str(len(date_results)),
        "With Flights": str(len(valid_prices)),
        "Cheapest Date": f"{cheapest['date']} ({cheapest['day']})",
        "Cheapest Price": f"{cheapest['price']} — {cheapest['airline']}",
        "Average Price": f"${price_avg}" if price_avg else "N/A",
        "Price Range": f"${price_min} – ${price_max}" if valid_prices else "N/A",
        "Source": "Google Flights via TARS v6.0",
    }
    for i, (k, v) in enumerate(summary.items()):
        ws.cell(row=srow + 1 + i, column=1, value=k).font = Font(name="Helvetica Neue", size=10, bold=True)
        ws.cell(row=srow + 1 + i, column=2, value=v).font = data_font

    widths = [5, 12, 12, 10, 18, 12, 14, 12, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════════════════
    #  SHEET 3: Insights (airline breakdown)
    # ═══════════════════════════════════════════════════
    if airline_counts:
        ws2 = wb.create_sheet(title="Insights")
        ws2.sheet_properties.tabColor = "F59E0B"
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws2.cell(row=1, column=1, value="💡 Cheapest Dates Intelligence").font = Font(
            name="Helvetica Neue", size=16, bold=True, color=DARK)
        ws2.row_dimensions[1].height = 36

        ws2.cell(row=3, column=1, value="📊 PRICE ANALYTICS").font = Font(
            name="Helvetica Neue", size=12, bold=True, color=DARK)
        stat_items = [
            ("Dates Scanned", len(date_results)),
            ("With Flights", len(valid_prices)),
            ("Cheapest", f"${price_min}"),
            ("Average", f"${price_avg}"),
            ("Highest", f"${price_max}"),
            ("Range", f"${price_max - price_min}" if valid_prices else "N/A"),
            ("Airlines Found", len(airline_counts)),
        ]
        for i, (label, val) in enumerate(stat_items):
            ws2.cell(row=4 + i, column=1, value=label).font = Font(name="Helvetica Neue", size=10, bold=True)
            ws2.cell(row=4 + i, column=2, value=str(val)).font = data_font

        # Airline Breakdown Table
        arow = 4 + len(stat_items) + 2
        ws2.cell(row=arow, column=1, value="✈️ AIRLINE COMPARISON").font = Font(
            name="Helvetica Neue", size=12, bold=True, color=DARK)
        arow += 1
        for col_idx, h in enumerate(["Airline", "Cheapest", "Average", "Dates Available", "Status"], 1):
            cell = ws2.cell(row=arow, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        sorted_airlines = sorted(airline_counts.items(), key=lambda x: x[1]["min_price"])
        for ai, (name, data) in enumerate(sorted_airlines):
            r = arow + 1 + ai
            ws2.cell(row=r, column=1, value=name).font = Font(name="Helvetica Neue", size=10, bold=True)
            ws2.cell(row=r, column=2, value=data["min_price"]).font = data_font
            ws2.cell(row=r, column=2).number_format = '$#,##0'
            ws2.cell(row=r, column=3, value=data["avg_price"]).font = data_font
            ws2.cell(row=r, column=3).number_format = '$#,##0'
            ws2.cell(row=r, column=4, value=data["count"]).font = data_font
            if ai == 0:
                ws2.cell(row=r, column=5, value="💰 Best Price").font = Font(
                    name="Helvetica Neue", size=10, bold=True, color=GREEN)
            elif data["min_price"] <= price_avg:
                ws2.cell(row=r, column=5, value="Competitive").font = Font(
                    name="Helvetica Neue", size=10, color=BLUE)
            else:
                ws2.cell(row=r, column=5, value="Above Average").font = Font(
                    name="Helvetica Neue", size=10, color=RED)

        # Data bars on Cheapest column
        cheapest_range = f"B{arow+1}:B{arow+len(sorted_airlines)}"
        ws2.conditional_formatting.add(cheapest_range, DataBarRule(
            start_type="min", end_type="max",
            color="059669", showValue=True, minLength=None, maxLength=None,
        ))

        # ── Smart Tips ──
        tips_row = arow + len(sorted_airlines) + 3
        ws2.cell(row=tips_row, column=1, value="💡 SMART TIPS").font = Font(
            name="Helvetica Neue", size=12, bold=True, color=DARK)
        tips = []
        if len(valid_prices) >= 3:
            cheapest_day = cheapest.get("day", "")
            tips.append(f"🏆 Best day to fly: {cheapest_day} — {cheapest['price']} on {cheapest['airline']}")
        if price_max > price_min * 2:
            tips.append(f"📈 Prices vary by {round((price_max/price_min - 1)*100)}% — date flexibility saves big!")
        if len(airline_counts) > 1:
            tips.append(f"✈️ {len(airline_counts)} airlines compete on this route — compare for best deals")
        cheapest_airline = sorted_airlines[0][0] if sorted_airlines else ""
        if cheapest_airline:
            tips.append(f"💰 {cheapest_airline} consistently offers the lowest prices")
        tips.append(f"📅 Scanned {len(date_results)} dates across {(end - start).days} day range")
        for ti, tip in enumerate(tips):
            ws2.cell(row=tips_row + 1 + ti, column=1, value=tip).font = data_font

        for col in range(1, 6):
            ws2.column_dimensions[get_column_letter(col)].width = 20

    safe = f"Cheapest_Dates_{origin_code}_{dest_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(REPORT_DIR, safe)
    wb.save(filepath)
    return {"success": True, "path": filepath}


# ═══════════════════════════════════════════════════════
#  PHASE 6 — Send HTML Email via SMTP (Outlook Direct)
# ═══════════════════════════════════════════════════════

def _send_html_email(to_address, subject, html_body, attachment_path="", from_address="tarsitgroup@outlook.com"):
    """Send a rich HTML email via Outlook SMTP directly.

    Uses Python smtplib + email.mime for guaranteed HTML rendering.
    No more Mail.app AppleScript — this ensures the HTML actually shows
    with charts, colors, and formatting in every email client.

    Outlook SMTP: smtp-mail.outlook.com:587 (STARTTLS)
    """
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    SMTP_SERVER = "smtp-mail.outlook.com"
    SMTP_PORT = 587
    SMTP_USER = from_address
    # App password for Outlook SMTP — stored here for simplicity
    SMTP_PASS = _get_smtp_password()

    if not SMTP_PASS:
        # Fallback to Mail.app AppleScript if no SMTP password configured
        return _send_html_email_mailapp(to_address, subject, html_body, attachment_path, from_address)

    try:
        msg = MIMEMultipart("mixed")
        msg["From"] = f"TARS Flight Intel <{from_address}>"
        msg["To"] = to_address
        msg["Subject"] = subject
        msg["X-Mailer"] = "TARS Flight Intelligence v5.0"

        # Attach HTML body — this is the key: MIMEText with 'html' subtype
        html_part = MIMEText(html_body, "html", "utf-8")
        msg.attach(html_part)

        # Attach Excel file if provided
        if attachment_path and os.path.isfile(attachment_path):
            with open(attachment_path, "rb") as f:
                part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                filename = os.path.basename(attachment_path)
                part.add_header("Content-Disposition", f"attachment; filename=\"{filename}\"")
                msg.attach(part)

        # Send via Outlook SMTP
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(from_address, to_address, msg.as_string())

        return {"success": True, "content": f"HTML email sent to {to_address}: {subject}"}

    except smtplib.SMTPAuthenticationError:
        logger.warning("    ⚠️ SMTP auth failed — falling back to Mail.app")
        return _send_html_email_mailapp(to_address, subject, html_body, attachment_path, from_address)
    except Exception as e:
        logger.warning(f"    ⚠️ SMTP failed ({e}) — falling back to Mail.app")
        return _send_html_email_mailapp(to_address, subject, html_body, attachment_path, from_address)


def _get_smtp_password():
    """Load SMTP password from config or environment."""
    # Try environment variable first
    pwd = os.environ.get("TARS_SMTP_PASSWORD", "")
    if pwd:
        return pwd
    # Try config.yaml
    try:
        config = _load_config()
        return config.get("email", {}).get("smtp_password", "")
    except Exception:
        return ""


def _send_html_email_mailapp(to_address, subject, html_body, attachment_path="", from_address="tarsitgroup@outlook.com"):
    """Send HTML email via Microsoft Outlook (preferred) or Mail.app (fallback).
    
    Outlook handles HTML content + attachments correctly.
    Mail.app has a known bug where adding attachments resets HTML body.
    """
    import subprocess
    import tempfile

    html_file = tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False)
    html_file.write(html_body)
    html_file.close()

    safe_subject = subject.replace('\\', '\\\\').replace('"', '\\"')

    # Try Microsoft Outlook first (handles HTML + attachments properly)
    outlook_result = _send_via_outlook(to_address, safe_subject, html_file.name, attachment_path)
    if outlook_result.get("success"):
        try: os.unlink(html_file.name)
        except Exception: pass
        return outlook_result

    # Fallback: Mail.app (set HTML AFTER attachment to avoid body reset bug)
    logger.warning(f"    ⚠️ Outlook failed, trying Mail.app...")
    if attachment_path and os.path.isfile(attachment_path):
        script = f'''
        set htmlContent to read POSIX file "{html_file.name}" as «class utf8»
        tell application "Mail"
            set msg to make new outgoing message with properties {{subject:"{safe_subject}", visible:false}}
            tell msg
                make new to recipient at end of to recipients with properties {{address:"{to_address}"}}
                set theAttachment to POSIX file "{attachment_path}"
                make new attachment with properties {{file name:theAttachment}} at after last paragraph
                delay 1
                set html content to htmlContent
            end tell
            delay 1
            send msg
        end tell
        '''
    else:
        script = f'''
        set htmlContent to read POSIX file "{html_file.name}" as «class utf8»
        tell application "Mail"
            set msg to make new outgoing message with properties {{subject:"{safe_subject}", visible:false}}
            tell msg
                set html content to htmlContent
                make new to recipient at end of to recipients with properties {{address:"{to_address}"}}
            end tell
            send msg
        end tell
        '''

    try:
        result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=60)
        os.unlink(html_file.name)
        if result.returncode == 0:
            return {"success": True, "content": f"HTML email sent via Mail.app to {to_address}: {subject}"}
        else:
            return {"success": False, "content": f"Mail.app error: {result.stderr}"}
    except Exception as e:
        try: os.unlink(html_file.name)
        except Exception: pass
        return {"success": False, "content": f"Email failed: {e}"}


def _send_via_outlook(to_address, subject, html_file_path, attachment_path=""):
    """Send HTML email via Microsoft Outlook AppleScript.
    
    Outlook properly renders HTML body AND supports attachments without
    stripping the HTML content (unlike Mail.app).
    """
    import subprocess

    attachment_block = ""
    if attachment_path and os.path.isfile(attachment_path):
        attachment_block = f'''
                make new attachment with properties {{file:POSIX file "{attachment_path}"}}
        '''

    script = f'''
    set htmlContent to read POSIX file "{html_file_path}" as «class utf8»
    tell application "Microsoft Outlook"
        set newMsg to make new outgoing message with properties {{subject:"{subject}", content:htmlContent}}
        tell newMsg
            make new to recipient with properties {{email address:{{address:"{to_address}"}}}}
            {attachment_block}
        end tell
        send newMsg
    end tell
    '''

    try:
        result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            return {"success": True, "content": f"HTML email sent via Outlook to {to_address}: {subject}"}
        else:
            logger.warning(f"    ⚠️ Outlook error: {result.stderr[:200]}")
            return {"success": False, "content": f"Outlook error: {result.stderr}"}
    except Exception as e:
        return {"success": False, "content": f"Outlook failed: {e}"}


# ═══════════════════════════════════════════════════════
#  PHASE 9 — Flight Report Pipeline
# ═══════════════════════════════════════════════════════

def search_flights_report(
    origin: str, destination: str, depart_date: str,
    return_date: str = "", passengers: int = 1,
    trip_type: str = "round_trip", cabin: str = "economy",
    stops: str = "any", sort_by: str = "price",
    max_price: int = 0, email_to: str = "",
) -> dict:
    """Search flights → elegant Excel → HTML email → iMessage notification."""
    result = search_flights(
        origin=origin, destination=destination, depart_date=depart_date,
        return_date=return_date, passengers=passengers, trip_type=trip_type,
        cabin=cabin, stops=stops, sort_by=sort_by, max_price=max_price,
    )
    flights = result.get("flights", [])
    if not result.get("success") or not flights:
        return result

    origin_code = _resolve_airport(origin)
    dest_code = _resolve_airport(destination)
    depart = _parse_date(depart_date)
    ret_parsed = _parse_date(return_date) if return_date else ""
    search_url = result.get("url", "")
    cheapest = flights[0]
    nonstops = [f for f in flights if "nonstop" in f.get("stops", "").lower()]
    best_ns = nonstops[0] if nonstops else None

    # v4.0/v5.0 — Grab intelligence from search result
    analytics = result.get("analytics", {})
    suggestions = result.get("suggestions", [])
    price_insight = result.get("price_insight", "")
    return_flight = result.get("return_flight", {})
    tracker_suggestion = result.get("tracker_suggestion", "")

    summary = {
        "Route": f"{origin_code} → {dest_code}", "Departure": depart,
        "Return": ret_parsed or "One-way", "Total Options": str(len(flights)),
        "Cheapest": f"{cheapest.get('price', '?')} — {cheapest.get('airline', '?')} ({cheapest.get('stops', '?')})",
    }
    if best_ns:
        summary["Cheapest Nonstop"] = f"{best_ns.get('price', '?')} — {best_ns.get('airline', '?')}"
    if analytics.get("price_avg"):
        summary["Average Price"] = f"${analytics['price_avg']}"
    if price_insight:
        summary["Google Insight"] = price_insight
    summary["Search Date"] = datetime.now().strftime("%B %d, %Y")
    summary["Source"] = "Google Flights"

    title = f"Flights {origin_code}→{dest_code} {depart}"
    excel_result = _generate_flight_excel(title, flights, origin_code, dest_code, search_url, summary,
                                           analytics=analytics, suggestions=suggestions)
    excel_path = excel_result.get("path", "")

    emailed = False
    email_msg = ""
    if email_to:
        try:
            html = _html_flight_report_email(origin_code, dest_code, depart, ret_parsed, flights, search_url,
                                              price_insight=price_insight, return_flight=return_flight,
                                              tracker_suggestion=tracker_suggestion)
            mail_result = _send_html_email(
                to_address=email_to,
                subject=f"✈️ {origin_code}→{dest_code} Flight Report · {cheapest.get('price', '')} cheapest",
                html_body=html, attachment_path=excel_path,
            )
            if mail_result.get("success"):
                emailed = True
                email_msg = f"\n📧 Report emailed to {email_to}"
            else:
                email_msg = f"\n⚠️ Email failed: {mail_result.get('content', '')}"
        except Exception as e:
            email_msg = f"\n⚠️ Email failed: {e}"

    content = result["content"] + f"\n\n📊 Excel saved: {excel_path}{email_msg}"

    try:
        from voice.imessage_send import IMessageSender
        sender = IMessageSender(_load_config())
        nonstop_line = ""
        if nonstops:
            nonstop_line = f"\n✈️ Best nonstop: {best_ns.get('price', '?')} — {best_ns.get('airline', '?')} ({best_ns.get('duration', '?')})"
        # Use airline-specific link for cheapest flight
        cheapest_link = cheapest.get("booking_link", search_url)

        # v4.0/v5.0 — Include top suggestions + tracker hint in iMessage
        tips_line = ""
        if suggestions:
            top_tips = suggestions[:2]
            tips_line = "\n\n💡 Tips:"
            for tip in top_tips:
                tips_line += f"\n{tip['icon']} {tip['text']}"

        insight_line = ""
        if price_insight:
            insight_line = f"\n📈 {price_insight}"

        tracker_line = ""
        if tracker_suggestion:
            tracker_line = f"\n{tracker_suggestion}"

        imsg = (
            f"✅ Flight report ready!\n\n"
            f"🛫 {origin_code} → {dest_code}\n"
            f"📅 {depart}" + (f" → {ret_parsed}" if ret_parsed else "") + "\n"
            f"📊 Found {len(flights)} options\n"
            f"💰 Cheapest: {cheapest.get('price', '?')} — {cheapest.get('airline', '?')} ({cheapest.get('stops', '?')})"
            f"{nonstop_line}"
            f"{insight_line}"
            f"{tips_line}"
            f"{tracker_line}\n\n"
            f"🔗 Book cheapest: {cheapest_link}"
        )
        if emailed:
            imsg += f"\n📧 Report emailed to {email_to}"
        sender.send(imsg)
    except Exception:
        pass

    return {"success": True, "content": content, "flights": flights,
            "excel_path": excel_path, "emailed": emailed, "url": search_url}


# ═══════════════════════════════════════════════════════
#  PHASE 3 — Cheapest Dates (Smart 6-Month Sampling)
# ═══════════════════════════════════════════════════════

def _search_single_date(args: tuple) -> dict:
    """Worker function for parallel cheapest dates scanning."""
    origin, destination, dt, trip_type, cabin, stops = args
    date_str = dt.strftime("%Y-%m-%d")
    try:
        r = search_flights(
            origin=origin, destination=destination,
            depart_date=date_str, trip_type=trip_type,
            cabin=cabin, stops=stops,
        )
        flights = r.get("flights", [])
        gf_link = _build_booking_link(origin, destination, date_str, trip_type=trip_type)
        if flights:
            best = flights[0]
            price = _price_num(best.get("price", "$99999"))
            verified_url = _get_verified_booking_link(
                best.get("airline", ""), origin, destination, date_str)
            booking_link = verified_url or best.get("booking_link", gf_link)
            result = {
                "date": date_str, "day": dt.strftime("%A"),
                "price": best.get("price", "—"), "price_num": price,
                "airline": best.get("airline", "—"), "stops": best.get("stops", "—"),
                "duration": best.get("duration", "—"), "depart_time": best.get("depart_time", "—"),
                "options": len(flights), "booking_link": booking_link,
                "price_insight": r.get("price_insight", ""),
            }
            logger.info(f"      📅 {date_str}: {best.get('price', '?')} ({best.get('airline', '?')})")
            return result
        else:
            logger.info(f"      📅 {date_str}: no results")
            return {
                "date": date_str, "day": dt.strftime("%A"),
                "price": "N/A", "price_num": 99999, "airline": "—", "stops": "—",
                "duration": "—", "depart_time": "—", "options": 0, "booking_link": gf_link,
                "price_insight": "",
            }
    except Exception as e:
        logger.info(f"      📅 {date_str}: error ({e})")
        return {
            "date": date_str, "day": dt.strftime("%A"),
            "price": "Error", "price_num": 99999, "airline": "—", "stops": "—",
            "duration": "—", "depart_time": "—", "options": 0,
            "booking_link": _build_booking_link(origin, destination, date_str, trip_type=trip_type),
            "price_insight": "",
        }


def find_cheapest_dates(
    origin: str, destination: str, start_date: str,
    end_date: str = "", trip_type: str = "one_way",
    cabin: str = "economy", stops: str = "any", email_to: str = "",
) -> dict:
    """Find the cheapest day to fly within a date range (up to 6 months).

    v5.0: Parallel scanning with ThreadPoolExecutor (2 workers to avoid
    overwhelming CDP). Cache integration means repeated dates are instant.

    Smart sampling:
      ≤14 days  → every day       (~14 searches)
      ≤30 days  → every 2 days    (~15 searches)
      ≤90 days  → every 5 days    (~18 searches)
      ≤180 days → every 7 days    (~26 searches)
      >180 days → every 10 days   (~25 searches)
    """
    origin_code = _resolve_airport(origin)
    dest_code = _resolve_airport(destination)

    start = datetime.strptime(_parse_date(start_date), "%Y-%m-%d")
    if end_date:
        end = datetime.strptime(_parse_date(end_date), "%Y-%m-%d")
    else:
        end = start + timedelta(days=30)

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    if start < today:
        start = today + timedelta(days=1)

    total_days = (end - start).days
    if total_days <= 0:
        return {"success": False, "content": "❌ end_date must be after start_date", "dates": []}

    if total_days <= 14:
        step = 1
    elif total_days <= 30:
        step = 2
    elif total_days <= 90:
        step = 5
    elif total_days <= 180:
        step = 7
    else:
        step = 10

    dates_to_search = []
    d = start
    while d <= end:
        dates_to_search.append(d)
        d += timedelta(days=step)
    if dates_to_search[-1] != end:
        dates_to_search.append(end)

    logger.info(f"    ✈️ Scanning {len(dates_to_search)} dates (parallel): {origin_code}→{dest_code} ({start.strftime('%b %d')} – {end.strftime('%b %d, %Y')})")

    # Build worker args
    worker_args = [(origin, destination, dt, trip_type, cabin, stops) for dt in dates_to_search]

    # Phase 14: Parallel scanning with 2 workers (safe for CDP)
    date_results = []
    with ThreadPoolExecutor(max_workers=2) as executor:
        futures = {executor.submit(_search_single_date, args): args[2] for args in worker_args}
        for future in as_completed(futures):
            try:
                result = future.result(timeout=120)
                if result:
                    date_results.append(result)
            except Exception as e:
                dt = futures[future]
                logger.info(f"      📅 {dt.strftime('%Y-%m-%d')}: worker error ({e})")
                date_results.append({
                    "date": dt.strftime("%Y-%m-%d"), "day": dt.strftime("%A"),
                    "price": "Error", "price_num": 99999, "airline": "—", "stops": "—",
                    "duration": "—", "depart_time": "—", "options": 0,
                    "booking_link": _build_booking_link(origin, destination, dt.strftime("%Y-%m-%d"), trip_type=trip_type),
                    "price_insight": "",
                })

    if not date_results:
        return {"success": False, "content": "❌ No results for any date in range", "dates": []}

    date_results.sort(key=lambda x: x["price_num"])
    cheapest = date_results[0]

    report_lines = [
        f"✈️ **Cheapest Dates: {origin_code} → {dest_code}**",
        f"📅 Range: {start.strftime('%b %d')} – {end.strftime('%b %d, %Y')}",
        f"Scanned {len(date_results)} dates\n",
        f"{'Date':<14} {'Day':<10} {'Price':<10} {'Airline':<15} {'Stops':<10} {'Duration'}",
        "─" * 75,
    ]
    for dr in date_results:
        report_lines.append(
            f"{dr['date']:<14} {dr['day']:<10} {dr['price']:<10} "
            f"{dr['airline']:<15} {dr['stops']:<10} {dr['duration']}"
        )
    report_lines.append(f"\n🏆 **CHEAPEST**: {cheapest['date']} ({cheapest['day']}) — "
                        f"{cheapest['price']} on {cheapest['airline']} "
                        f"({cheapest['stops']}, {cheapest['duration']})")
    if len(date_results) >= 3:
        report_lines.append("\n💡 Top 3 cheapest dates:")
        for i, dr in enumerate(date_results[:3], 1):
            report_lines.append(f"  {i}. {dr['date']} ({dr['day']}) — {dr['price']} ({dr['airline']}) — {dr.get('booking_link', '')}")
    report = "\n".join(report_lines)

    excel_result = _generate_dates_excel(origin_code, dest_code, date_results, start, end)
    excel_path = excel_result.get("path", "")
    if excel_path:
        report += f"\n\n📊 Excel saved: {excel_path}"

    emailed = False
    if email_to and excel_path:
        try:
            html = _html_cheapest_dates_email(
                origin_code, dest_code, date_results,
                start.strftime('%b %d'), end.strftime('%b %d, %Y'),
            )
            mail_result = _send_html_email(
                to_address=email_to,
                subject=f"📊 Cheapest Dates: {origin_code}→{dest_code} · {cheapest['price']} lowest",
                html_body=html, attachment_path=excel_path,
            )
            if mail_result.get("success"):
                emailed = True
                report += f"\n📧 Report emailed to {email_to}"
            else:
                report += f"\n⚠️ Email failed: {mail_result.get('content', '')}"
        except Exception as e:
            report += f"\n⚠️ Email failed: {e}"

    try:
        from voice.imessage_send import IMessageSender
        sender = IMessageSender(_load_config())
        imsg = (
            f"✅ Cheapest dates report ready!\n\n"
            f"🛫 {origin_code} → {dest_code}\n"
            f"📅 {start.strftime('%b %d')} – {end.strftime('%b %d, %Y')}\n"
            f"📊 Scanned {len(date_results)} dates\n\n"
            f"🏆 Cheapest: {cheapest['date']} ({cheapest['day']}) — "
            f"{cheapest['price']} on {cheapest['airline']}"
        )
        if len(date_results) >= 3:
            imsg += "\n\n💡 Top 3:"
            for i, dr in enumerate(date_results[:3], 1):
                imsg += f"\n  {i}. {dr['date']} — {dr['price']} ({dr['airline']})"
        imsg += f"\n\n🔗 Book cheapest: {cheapest.get('booking_link', '')}"
        if emailed:
            imsg += f"\n📧 Full report emailed to {email_to}"
        sender.send(imsg)
    except Exception:
        pass

    return {"success": True, "content": report, "dates": date_results,
            "cheapest": cheapest, "excel_path": excel_path, "emailed": emailed}


# ═══════════════════════════════════════════════════════
#  PHASE 6 & 7 — Price Tracker + Alert Engine
# ═══════════════════════════════════════════════════════

TRACKER_DB = os.path.expanduser("~/Documents/TARS_Reports/flight_trackers.json")


def _load_trackers():
    if os.path.isfile(TRACKER_DB):
        try:
            with open(TRACKER_DB, "r") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return []
    return []


def _save_trackers(trackers):
    os.makedirs(os.path.dirname(TRACKER_DB), exist_ok=True)
    with open(TRACKER_DB, "w") as f:
        json.dump(trackers, f, indent=2)


def track_flight_price(
    origin: str, destination: str, depart_date: str,
    target_price: int, return_date: str = "",
    trip_type: str = "round_trip", cabin: str = "economy",
    stops: str = "any", email_to: str = "tarsitgroup@outlook.com",
    check_interval_hours: int = 6,
) -> dict:
    """Set up a price tracker for a specific route.

    Monitors price on Google Flights at regular intervals.
    When price drops to or below target, sends:
      - Beautiful HTML email alert with booking link
      - iMessage notification
    """
    origin_code = _resolve_airport(origin)
    dest_code = _resolve_airport(destination)
    depart = _parse_date(depart_date)
    ret = _parse_date(return_date) if return_date else ""

    tracker_id = f"{origin_code}-{dest_code}-{depart[:10].replace('-', '')}"

    logger.info(f"    🎯 Setting up tracker: {origin_code}→{dest_code} on {depart}, target ${target_price}")
    initial = search_flights(
        origin=origin, destination=destination, depart_date=depart_date,
        return_date=return_date, trip_type=trip_type, cabin=cabin, stops=stops,
    )

    current_price = None
    current_airline = "—"
    if initial.get("flights"):
        best = initial["flights"][0]
        current_price = _price_num(best.get("price", "$99999"))
        current_airline = best.get("airline", "—")

    tracker = {
        "id": tracker_id, "origin": origin_code, "destination": dest_code,
        "depart_date": depart, "return_date": ret,
        "trip_type": trip_type, "cabin": cabin, "stops": stops,
        "target_price": target_price, "email_to": email_to,
        "check_interval_hours": check_interval_hours,
        "created_at": datetime.now().isoformat(),
        "last_checked": datetime.now().isoformat(),
        "last_price": current_price, "last_airline": current_airline,
        "price_history": [{"timestamp": datetime.now().isoformat(), "price": current_price, "airline": current_airline}] if current_price else [],
        "alerts_sent": 0, "active": True,
    }

    trackers = _load_trackers()
    trackers = [t for t in trackers if t["id"] != tracker_id]
    trackers.append(tracker)
    _save_trackers(trackers)

    already_hit = current_price and current_price <= target_price
    if already_hit:
        verified_link = _get_verified_booking_link(current_airline, origin, destination, depart_date, return_date)
        booking_link = verified_link or _build_booking_link(origin, destination, depart_date, return_date, trip_type)
        _send_price_alert(tracker, current_price, current_airline, booking_link)
        content = (
            f"🎯 Tracker set: {origin_code}→{dest_code} on {depart}\n"
            f"🔔 **ALERT: Already at target!** Current ${current_price} ≤ target ${target_price}\n"
            f"📧 Alert email sent immediately!\n"
            f"🔗 Book: {booking_link}"
        )
    else:
        content = (
            f"🎯 Tracker set: {origin_code}→{dest_code} on {depart}\n"
            f"💰 Target: ${target_price}\n"
            f"📊 Current price: ${current_price or 'unknown'} ({current_airline})\n"
            f"⏰ Checking every {check_interval_hours} hours\n"
            f"📧 Alert will be sent to {email_to}\n"
            f"🆔 Tracker ID: {tracker_id}"
        )

    try:
        from voice.imessage_send import IMessageSender
        sender = IMessageSender(_load_config())
        imsg = (
            f"🎯 Flight price tracker activated!\n\n"
            f"🛫 {origin_code} → {dest_code}\n"
            f"📅 {depart}\n"
            f"💰 Target: ${target_price}\n"
            f"📊 Current: ${current_price or '?'} ({current_airline})\n"
            f"⏰ Checking every {check_interval_hours}h\n"
            f"🆔 {tracker_id}"
        )
        if already_hit:
            imsg += f"\n\n🔔 ALREADY AT TARGET! Check your email!"
        sender.send(imsg)
    except Exception:
        pass

    return {"success": True, "tracker_id": tracker_id, "content": content,
            "current_price": current_price, "target_price": target_price,
            "already_at_target": already_hit}


def get_tracked_flights() -> dict:
    """Get all active flight price trackers with their status."""
    trackers = _load_trackers()
    active = [t for t in trackers if t.get("active")]
    if not active:
        return {"success": True, "content": "📭 No active flight trackers.\n\nSet one up with track_flight_price.", "trackers": []}

    lines = ["🎯 **Active Flight Price Trackers**\n"]
    lines.append(f"{'ID':<25} {'Route':<12} {'Date':<12} {'Target':<10} {'Last Price':<12} {'Status'}")
    lines.append("─" * 85)

    for t in active:
        route = f"{t['origin']}→{t['destination']}"
        last = t.get("last_price")
        target = t["target_price"]
        status = "✅ At target!" if (last and last <= target) else "⏳ Monitoring"
        lines.append(f"{t['id']:<25} {route:<12} {t['depart_date']:<12} ${target:<9} {'$' + str(last) if last else 'N/A':<11} {status}")
        if t.get("price_history"):
            prices = [p["price"] for p in t["price_history"] if p.get("price")]
            if len(prices) >= 2:
                trend = "📈" if prices[-1] > prices[-2] else "📉" if prices[-1] < prices[-2] else "➡️"
                lines.append(f"  {trend} Price trend: {', '.join(f'${p}' for p in prices[-5:])}")

    lines.append(f"\n📊 Total active trackers: {len(active)}")
    return {"success": True, "content": "\n".join(lines), "trackers": active}


def stop_tracking(tracker_id: str) -> dict:
    """Stop a specific flight price tracker."""
    trackers = _load_trackers()
    found = False
    for t in trackers:
        if t["id"] == tracker_id:
            t["active"] = False
            found = True
            break
    if found:
        _save_trackers(trackers)
        return {"success": True, "content": f"🛑 Tracker {tracker_id} stopped."}
    return {"success": False, "content": f"❌ Tracker {tracker_id} not found."}


def _send_price_alert(tracker, current_price, airline, booking_link):
    """Send a price alert via HTML email and iMessage."""
    origin, dest = tracker["origin"], tracker["destination"]
    target, depart = tracker["target_price"], tracker["depart_date"]
    email_to = tracker.get("email_to", "tarsitgroup@outlook.com")
    tracker_id = tracker["id"]

    try:
        html = _html_price_alert_email(origin, dest, target, current_price, airline, depart, booking_link, tracker_id)
        _send_html_email(to_address=email_to, subject=f"🔔 Price Alert: {origin}→{dest} dropped to ${current_price}!", html_body=html)
    except Exception as e:
        logger.warning(f"    ⚠️ Alert email failed: {e}")

    try:
        from voice.imessage_send import IMessageSender
        sender = IMessageSender(_load_config())
        savings = target - current_price if current_price < target else 0
        imsg = (
            f"🔔 PRICE ALERT!\n\n"
            f"🛫 {origin} → {dest}\n📅 {depart}\n"
            f"💰 Now: ${current_price} (target was ${target})\n✈️ {airline}"
        )
        if savings > 0:
            imsg += f"\n💚 You save ${savings}!"
        imsg += f"\n\n🔗 Book now: {booking_link}"
        sender.send(imsg)
    except Exception:
        pass


def check_price_trackers() -> dict:
    """Check all active trackers and send alerts for any that hit target.
    Called by the TARS scheduler (background thread).
    """
    trackers = _load_trackers()
    active = [t for t in trackers if t.get("active")]
    if not active:
        return {"checked": 0, "alerts": 0}

    now = datetime.now()
    checked = 0
    alerts = 0

    for tracker in active:
        last_checked = datetime.fromisoformat(tracker.get("last_checked", "2000-01-01"))
        interval = timedelta(hours=tracker.get("check_interval_hours", 6))
        if now - last_checked < interval:
            continue

        try:
            depart = datetime.strptime(tracker["depart_date"], "%Y-%m-%d")
            if depart < now:
                tracker["active"] = False
                continue
        except ValueError:
            pass

        logger.debug(f"    🔍 Checking tracker: {tracker['id']}...")
        checked += 1

        result = search_flights(
            origin=tracker["origin"], destination=tracker["destination"],
            depart_date=tracker["depart_date"], return_date=tracker.get("return_date", ""),
            trip_type=tracker.get("trip_type", "round_trip"),
            cabin=tracker.get("cabin", "economy"), stops=tracker.get("stops", "any"),
        )

        tracker["last_checked"] = now.isoformat()

        if result.get("flights"):
            best = result["flights"][0]
            price = _price_num(best.get("price", "$99999"))
            airline = best.get("airline", "—")
            tracker["last_price"] = price
            tracker["last_airline"] = airline
            tracker["price_history"].append({"timestamp": now.isoformat(), "price": price, "airline": airline})
            tracker["price_history"] = tracker["price_history"][-50:]

            if price <= tracker["target_price"]:
                # Use verified airline booking link for alerts
                verified_link = _get_verified_booking_link(
                    airline, tracker["origin"], tracker["destination"],
                    tracker["depart_date"], tracker.get("return_date", ""))
                booking_link = verified_link or _build_booking_link(
                    tracker["origin"], tracker["destination"],
                    tracker["depart_date"], tracker.get("return_date", ""),
                    tracker.get("trip_type", "round_trip"),
                )
                _send_price_alert(tracker, price, airline, booking_link)
                tracker["alerts_sent"] = tracker.get("alerts_sent", 0) + 1
                alerts += 1
                logger.info(f"    🔔 ALERT: {tracker['id']} at ${price} (target ${tracker['target_price']})")
            else:
                logger.info(f"    📊 {tracker['id']}: ${price} (target ${tracker['target_price']})")

    _save_trackers(trackers)
    return {"checked": checked, "alerts": alerts}


# ═══════════════════════════════════════════════════════
#  PHASE 7 — Background Scheduler Thread
# ═══════════════════════════════════════════════════════

_scheduler_thread = None
_scheduler_running = False


def start_price_tracker_scheduler(check_interval_minutes: int = 30):
    """Start the background scheduler that checks price trackers."""
    global _scheduler_thread, _scheduler_running
    if _scheduler_running:
        return
    _scheduler_running = True

    def _scheduler_loop():
        global _scheduler_running
        logger.info(f"    ⏰ Price tracker scheduler started (every {check_interval_minutes}min)")
        while _scheduler_running:
            try:
                result = check_price_trackers()
                if result["checked"] > 0:
                    logger.info(f"    ⏰ Tracker check: {result['checked']} checked, {result['alerts']} alerts")
            except Exception as e:
                logger.warning(f"    ⚠️ Tracker scheduler error: {e}")
            for _ in range(check_interval_minutes * 6):
                if not _scheduler_running:
                    break
                time.sleep(10)
        logger.info("    ⏰ Price tracker scheduler stopped")

    _scheduler_thread = threading.Thread(target=_scheduler_loop, daemon=True)
    _scheduler_thread.start()


def stop_price_tracker_scheduler():
    """Stop the background scheduler."""
    global _scheduler_running
    _scheduler_running = False


# ═══════════════════════════════════════════════════════
#  PHASE 8 — Direct Flight Booking via Browser
# ═══════════════════════════════════════════════════════

def book_flight(
    origin: str, destination: str, depart_date: str,
    return_date: str = "", airline: str = "",
    trip_type: str = "round_trip", cabin: str = "economy",
    passengers: int = 1, flight_number: int = 0,
) -> dict:
    """Book a flight by navigating the browser to the airline/Google Flights booking page.

    This uses CDP browser automation to:
      1. Search for the specific flight on Google Flights
      2. Click on the selected flight to open booking
      3. Navigate to the airline's checkout page
      4. Pre-fill passenger details where possible
      5. Hand off to the user for payment

    The function opens the booking flow in Chrome and guides through selection.
    User completes the final payment step.

    Args:
        origin: Departure city/airport code
        destination: Arrival city/airport code
        depart_date: Departure date
        return_date: Return date (optional)
        airline: Preferred airline name (optional - picks cheapest if empty)
        trip_type: round_trip or one_way
        cabin: economy, premium_economy, business, first
        passengers: Number of passengers
        flight_number: Index of flight to book from search results (0 = cheapest)
    """
    from hands.browser import act_goto, act_click, act_read_page, act_inspect_page, act_fill

    origin_code = _resolve_airport(origin)
    dest_code = _resolve_airport(destination)
    dep = _parse_date(depart_date) if not re.match(r'^\d{4}-\d{2}-\d{2}$', depart_date) else depart_date
    ret = ""
    if return_date:
        ret = _parse_date(return_date) if not re.match(r'^\d{4}-\d{2}-\d{2}$', return_date) else return_date

    logger.info(f"    🛒 Starting booking flow: {origin_code}→{dest_code} on {dep}")

    # Step 1: If airline is specified, try their direct booking page first
    if airline and airline != "—":
        airline_url = _get_airline_booking_url(airline, origin_code, dest_code, dep, ret)
        if airline_url and "google.com" not in airline_url:
            logger.info(f"    🌐 Opening {airline} booking page...")
            try:
                goto_result = act_goto(airline_url)
                time.sleep(3)
                page_text = act_read_page()
                
                # Notify user via iMessage
                try:
                    from voice.imessage_send import IMessageSender
                    sender = IMessageSender(_load_config())
                    sender.send(
                        f"🛒 Booking started!\n\n"
                        f"✈️ {origin_code} → {dest_code}\n"
                        f"📅 {dep}" + (f" → {ret}" if ret else "") + f"\n"
                        f"🏢 {airline}\n"
                        f"🌐 I've opened {airline}'s booking page in Chrome.\n\n"
                        f"👉 Check your Chrome browser to continue with passenger details and payment.\n\n"
                        f"🔗 {airline_url[:80]}"
                    )
                except Exception:
                    pass

                return {
                    "success": True,
                    "content": (
                        f"🛒 Booking page opened for {airline}!\n\n"
                        f"✈️ {origin_code} → {dest_code} on {dep}\n"
                        f"🌐 URL: {airline_url}\n\n"
                        f"Chrome is now showing {airline}'s booking page with your route pre-filled.\n"
                        f"Please check your Chrome browser to:\n"
                        f"  1. Select your preferred flight\n"
                        f"  2. Enter passenger details\n"
                        f"  3. Complete payment\n\n"
                        f"💡 I'll stay available if you need help during the process."
                    ),
                    "booking_url": airline_url,
                    "airline": airline,
                    "method": "airline_direct",
                }
            except Exception as e:
                logger.warning(f"    ⚠️ Airline booking page failed: {e}")

    # Step 2: Search on Google Flights first, then click through to book
    logger.debug(f"    🔍 Searching flights to find the best option to book...")

    # Search flights to find the right one
    search_result = search_flights(
        origin=origin_code, destination=dest_code,
        depart_date=dep, return_date=ret,
        trip_type=trip_type, cabin=cabin,
        passengers=passengers,
    )

    flights = search_result.get("flights", [])
    if not flights:
        return {
            "success": False,
            "content": f"❌ No flights found for {origin_code}→{dest_code} on {dep}. Cannot proceed with booking.",
        }

    # Select the right flight
    selected = None
    if airline and airline != "—":
        # Find a flight matching the preferred airline
        for f in flights:
            if airline.lower() in f.get("airline", "").lower():
                selected = f
                break
    
    if not selected:
        # Use flight_number index or cheapest
        idx = min(flight_number, len(flights) - 1)
        selected = flights[idx]

    sel_airline = selected.get("airline", "Unknown")
    sel_price = selected.get("price", "?")
    sel_depart = selected.get("depart_time", "?")
    sel_duration = selected.get("duration", "?")
    sel_stops = selected.get("stops", "?")

    logger.info(f"    ✈️ Selected: {sel_airline} at {sel_price} ({sel_stops}, {sel_duration})")

    # Step 3: Navigate to booking URL
    # Try airline-specific URL first
    booking_url = _get_verified_booking_link(sel_airline, origin_code, dest_code, dep, ret)
    
    logger.info(f"    🌐 Opening booking page: {booking_url[:60]}...")
    try:
        act_goto(booking_url)
        time.sleep(4)

        # Try to click "Book" or "Select" on Google Flights if we're there
        if "google.com/travel" in booking_url:
            try:
                # On Google Flights, click the matching flight result
                page_info = act_inspect_page()
                # Try clicking a "Select flight" or price button
                act_click(sel_price.replace("$", ""))
                time.sleep(2)
                # Look for "Book" button 
                act_click("Book")
                time.sleep(3)
                # Read the page to see if we landed on airline site
                final_page = act_read_page()
                logger.info(f"    📄 Landed on: {final_page[:100] if final_page else 'unknown'}...")
            except Exception as click_err:
                logger.warning(f"    ⚠️ Auto-click: {click_err}")

        # Notify user
        try:
            from voice.imessage_send import IMessageSender
            sender = IMessageSender(_load_config())
            sender.send(
                f"🛒 Booking in progress!\n\n"
                f"✈️ {origin_code} → {dest_code}\n"
                f"📅 {dep}" + (f" → {ret}" if ret else "") + f"\n"
                f"🏢 {sel_airline} — {sel_price}\n"
                f"⏰ {sel_depart} · {sel_duration} · {sel_stops}\n\n"
                f"🌐 Chrome is open on the booking page.\n"
                f"👉 Please complete passenger details and payment.\n\n"
                f"💡 Tell me if you need help with anything!"
            )
        except Exception:
            pass

        return {
            "success": True,
            "content": (
                f"🛒 Booking page opened!\n\n"
                f"Selected flight:\n"
                f"  ✈️ {sel_airline} — {sel_price}\n"
                f"  📅 {dep} at {sel_depart}\n"
                f"  ⏱️ {sel_duration} · {sel_stops}\n\n"
                f"🌐 Chrome is now showing the booking page.\n"
                f"Please check your browser to:\n"
                f"  1. Review flight details\n"
                f"  2. Enter passenger information\n"
                f"  3. Complete payment\n\n"
                f"🔗 {booking_url}\n\n"
                f"💡 I can also help fill in passenger details if you tell me your name and passport info."
            ),
            "booking_url": booking_url,
            "airline": sel_airline,
            "price": sel_price,
            "flight_details": selected,
            "method": "google_flights_redirect",
        }

    except Exception as e:
        # If browser automation fails, give them the direct link
        return {
            "success": True,
            "content": (
                f"🛒 Here's your booking link!\n\n"
                f"Selected: {sel_airline} — {sel_price} ({sel_stops})\n"
                f"📅 {dep} at {sel_depart}\n\n"
                f"🔗 Book here: {booking_url}\n\n"
                f"I couldn't auto-open the browser ({e}), but click the link above to book directly."
            ),
            "booking_url": booking_url,
            "airline": sel_airline,
            "price": sel_price,
            "method": "link_only",
        }

